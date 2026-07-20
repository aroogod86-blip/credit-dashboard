"""
Bloomberg IG 크레딧 채권 → 공유용 HTML 대시보드 생성기
======================================================
필요 패키지:  pip install xbbg pandas numpy

사용법:
  1. Bloomberg Terminal이 실행 중인 PC에서 실행
  2. python export_credit_dashboard.py
  3. 생성된 credit_dashboard_YYYYMMDD.html 파일을 팀에 공유

기존 bloomberg_credit_dashboard.py의 데이터 로드 로직을 그대로 활용하고,
Dash 서버 대신 standalone HTML 파일을 생성합니다.
"""

import sys
import io
# Windows CP949 환경에서 이모지/한글 출력 깨짐 방지
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if sys.stderr.encoding and sys.stderr.encoding.lower() not in ("utf-8", "utf8"):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import json
import os
import warnings
warnings.filterwarnings("ignore")

# ══════════════════════════════════════════════
# 1. ISIN 목록
# ══════════════════════════════════════════════
# bond_list.json이 있으면 거기서 읽고, 없으면 아래 기본 목록 사용
_DEFAULT_BOND_ISINS = [
    "US500630ED65",  "US50064FAY07",  "US50064FBA12",  "US037833FA32",
    "US302154EB15",  "USY8085FBL32",  "US02079KAX54",  "US89115A3C46",
    "USH4209UAC02",  "XS3219365635",  "XS2747557416",  "US02079KBM80",
    "US46647PEU66",  "US61690DK726",  "US86562MEJ53",  "USY8085FBD16",
    "USY8085FBY52",  "USY8085FBZ28",  "XS3030377132",  "USY8085FBK58",
    "US404280CF48",  "USY4899GGX52",  "USY8085FBU31",  "US46647PEY88",
    "US46647PEV40",  "USJ5S39RAQ78",  "US05964HAS40",  "US95000U3T82",
    "US46647PFD33",  "USJ7771YTM95",  "US404280EF20",  "US95000U3P60",
    "USJ5S39RAM64",  "US61747YFY68",  "US500631AE67",  "USY29011DM51",
    "USG84228FY98",  "US06051GHZ54",  "USY29011DG83",  "US46647PEJ12",
    "USJ5S39RAS35",  "US404280EM70",  "USH3698DBM59",  "US44891CDB63",
    "US65535HBF55",  "US06051GMY25",  "USY3815NBK64",  "US98105HAG56",
    "US780097BG51",  "USG84228FU76",  "US06738ECX13",  "XS3078370726",
    "US404280CV97",  "US38141GC365",  "USG84228FL77",  "USH42097EV54",
    "US404280BT50",  "US60687YDL02",  "US44891CDL46",  "USY7S272AG74",
    "XS1795263281",  "US65535HAY53",  "XS3109629371",  "US61748UAR32",
    "US780097BL47",  "USH42097FS17",  "US65535HBM07",  "XS2703610050",
    "USY52758AE20",  "US05964HBJ32",  "US06051GNA30",  "USY7S272AL69",
    "XS3187679041",  "US404280ER67",  "US44891CDV28",  "USY70750CB13",
    "US404280CC17",  "USH3698DCW23",  "US44891CDZ32",  "US82460EAN04",
    "USY5S80VAB27",  "US44891CDM29",  "US50050HAN61",  "XS2798085416",
    "US44891CCE12",  "USY5S5CGAK82",  "US44891CED11",  "USG84228FZ63",
    "XS3299373996",  "US09659X3B68",  "US44891CCU53",  "US46647PFJ03",
    "US46647PFE16",  "US44891CDQ33",  "US404280EN53",  "XS1932879130",
    "US86562MED83",  "XS2739009855",  "US09659X2X97",  "US606822DS05",
    "US44891CBL63",  "XS3189630372",  "USJ5903AAA28",  "US61748UAS15",
    "US44891CDW01",  "US44891CCJ09",  "USY5S5CGAN22",  "XS1452410571",
    "US83368TCJ51",  "US44891CCP68",  "USY5S5CGAB83",  "US44891CEA71",
    "USY5S5CGAL65",  "XS2985211569",  "US44891CEE93",  "US404280EG03",
    "US44891CDH34",  "USH42097FT99",  "US44891CDD20",  "USY7S272AM43",
    "USJ5S39RAU80",  "US65535HCB33",  "USJ5901UAH52",  "USY5S5CGAP79",
    "USY5S5CGAR36",  "KR6065904G23",  "XS2395327641",  "XS3017043053",
    "XS2559683359",  "XS3028206350",  "FR0014010FH7",
]

def load_bond_list():
    """bond_list.json이 있으면 거기서 ISIN 목록을 읽고, 없으면 기본 목록 사용"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    bond_list_path = os.path.join(script_dir, "bond_list.json")
    if os.path.exists(bond_list_path):
        try:
            with open(bond_list_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            isins = data if isinstance(data, list) else data.get("isins", [])
            if isins:
                print(f"  📋 bond_list.json에서 {len(isins)}개 종목 로드")
                return isins
        except Exception as e:
            print(f"  ⚠️  bond_list.json 읽기 오류: {e}")
    return _DEFAULT_BOND_ISINS

BOND_ISINS = load_bond_list()

ISSUER_CURVE_ISINS = [
    "US500630EG96", "US500630ED65", "US500630EH79",
    "US500630DP05", "US500630DU99", "US500630EB00",
    "US302154EA32", "US302154EQ83", "US302154EK14",
    "US302154ER66", "US302154DW60", "US302154ES40",
    "US44891CCX92", "US44891CBP77", "US44891CCD39",
    "US44891CDZ32", "US44891CCY75", "US44891CDG50",
    "US44891CDQ33", "US44891CEA71", "US44891CEE93",
    "US44891CDH34", "US44891CDX83",
    "USY8085FBT67", "USY8085FBY52", "USY8085FBZ28",
    "USY8085FBD16", "USY8085FBL32",
    "USY5S5CGAK82", "USY5S5CGAN22", "USY5S5CGAB83",
    "USY5S5CGAL65", "USY5S5CGAP79", "USY5S5CGAM49",
    "USY5S5CGAR36",
    "USY7S272AE27", "USY7S272AG74", "USY70750CB13",
    "USY7S272AL69", "USY7S272AH57", "USY70750CC95",
    "USY7S272AM43",
    "XS2861753924", "XS3189630372", "XS3299373996",
    "XS3299380850",
    "US61761JZN26", "XS2595028536", "US61748UAR32",
    "US61747YGB56", "US61748UAT97", "US61747YGC30",
    "US46625HJZ47", "US46625HRY89", "US46647PFD33",
    "US46647PFG63", "US46647PDH64", "XS2791972248",
    "US46647PFE16", "US46647PFJ03",
    "US404280DG12", "US404280DU06", "US404280FK06",
    "US404280CF48", "US404280FE46", "US404280FL88",
    "US404280DH94", "US404280DV88", "US404280FG93",
    "US02079KAW71", "US02079KBK25", "US02079KBL08",
    "US02079KBM80", "US02079KBN63", "US02079KBP12",
    "US023135DC78", "US023135DD51", "US023135DE35",
    "US023135DF00", "US023135DG82", "US023135DH65",
    "US68389XDX03", "US68389XDY85", "US68389XDM48",
    "US68389XDZ50", "US68389XCK90", "US68389XDR35",
    "US68389XEB73", "US68389XEC56",
    "USJ5S39RAL81", "USJ5S39RAD65", "USJ5S39RAM64",
    "USJ5S39RAS35", "USJ5S39RAE49", "USJ5S39RAU80",
    "USJ5S39RAV63",
    # ── 추가 발행자 곡선 종목 ──
    "ZO621460", "ZK896516", "ZF990477", "YO128804", "BP054719",
    "YS422819", "DC822725", "ZF990475", "YV600775", "ZH098092",
    "ZF097197", "YO680597", "DC696922", "ZJ311931", "YS344475",
    "DC696917", "ZI782661", "ZD475756", "AS592157", "DF616766",
    "YV564962", "BK129468", "BO389986", "ZN017996", "ZK471433",
    "YV564963", "YR685574", "BP135680", "BU068258", "ZN225587",
    "YO270397", "DI841976", "YS765590", "YI718036", "DI841979",
    "BU068259", "ZK292754", "ZF435445", "YO270399", "YP652772",
    "YI082278", "YJ138048", "DI424234", "YP652776", "YI082280",
    "BT631348", "DI424237", "YJ138052", "YI082283", "BW306958",
    "ZI966557", "YK266716", "YI794680", "ZB975719", "YK267609",
    "YI794697", "YJ946688", "ZI966562", "ZF069836", "YO269253",
    "YI794696", "ZB470534", "YP651966", "YL980741", "YI082274",
    "DK712152", "BX921089", "ZJ884878", "ZF374824", "YP651972",
    "YI082276",
    # ── AAPL / GOOGL / ORCL ──
    "YI913236", "YI913237", "YI913238", "YI913239", "YI913240",
    "DG517944", "DG517946", "DG517949", "DG517950", "DG517952",
    "DG517954",
    "YI663394", "YI663395", "YK411689", "YI663396", "ZN222552",
    "YK411691", "YI663398", "YI663399",
    "YX948674", "BO248052", "YX948675", "YM082834", "BO248054",
    "YM082836", "YM082837",
]

ALL_ISINS = list(dict.fromkeys(BOND_ISINS + ISSUER_CURVE_ISINS))

FIELDS = {
    "name":     "SECURITY_DES",
    "issuer":   "ISSUER",
    "rating":   "RTG_SP",
    "sector":   "BICS_LEVEL_1_SECTOR_NAME",
    "industry": "INDUSTRY_SECTOR",
    "maturity": "MATURITY",
    "coupon":   "CPN",
    "ytm":      "YLD_YTM_MID",
    "oas":      "OAS_SPREAD_MID",
    "spread":   "YAS_YLD_SPREAD",
    "zspread":  "Z_SPREAD_MID",
    "duration": "DUR_MID",
    "price":    "PX_LAST",
    "amt_out":  "AMT_OUTSTANDING",
}

UST_TICKERS = {
    "1M":"GB1 Govt","3M":"GB3 Govt","6M":"GB6 Govt",
    "1Y":"GB12 Govt","2Y":"GT2 Govt","3Y":"GT3 Govt",
    "5Y":"GT5 Govt","7Y":"GT7 Govt","10Y":"GT10 Govt",
    "20Y":"GT20 Govt","30Y":"GT30 Govt",
}

MARKET_TICKERS = {
    "LUACTRUU Index":     ("US IG OAS",  "INDEX_OAS_TSY_BP"),
    "LF98TRUU Index":     ("US HY OAS",  "INDEX_OAS_TSY_BP"),
    "CDX IG CDSI Curncy": ("CDX IG 5Y",  "PX_LAST"),
    "CDX HY CDSI Curncy": ("CDX HY 5Y",  "PX_LAST"),
    "VIX Index":          ("VIX",        "PX_LAST"),
    "MOVE Index":         ("MOVE",       "PX_LAST"),
}

ETF_TICKERS = {
    "VCSH US Equity":  "VCSH",
    "VCIT US Equity":  "VCIT",
    "LQD US Equity":   "LQD",
    "VCLT US Equity":  "VCLT",
    "IEAC LN Equity":  "IEAC",
    "HYG US Equity":   "HYG",
    "EMB US Equity":   "EMB",
    "EMLC US Equity":  "EMLC",
}

# ══════════════════════════════════════════════
# BDC 티커 정의
# ══════════════════════════════════════════════
BDC_TICKERS = {
    "ARCC US Equity": {"name": "Ares Capital",        "short": "ARCC",  "color": "#1B3A6B"},
    "OBDC US Equity": {"name": "Blue Owl (OBDC)",     "short": "OBDC",  "color": "#2E5FA3"},
    "FSK US Equity":  {"name": "FS KKR Capital",      "short": "FSK",   "color": "#D35400"},
    "BXSL US Equity": {"name": "Blackstone SL",       "short": "BXSL",  "color": "#1A6B3C"},
    "GBDC US Equity": {"name": "Golub Capital",       "short": "GBDC",  "color": "#C0392B"},
    "OTF US Equity":  {"name": "Blue Owl Tech (OTF)", "short": "OTF",   "color": "#8E44AD"},
}

END_DATE   = datetime.today()
START_DATE = END_DATE - timedelta(days=90)
MKT_START  = END_DATE - timedelta(days=180)

# Bloomberg 워크시트 경로 (PREL - 최근 발행)
PREL_FILE_PATH = r"C:\blp\data"
PREL_FILE_PATTERN = "PREL"


# ══════════════════════════════════════════════
# 1.5 PREL 워크시트 로드
# ══════════════════════════════════════════════
def load_prel_data(folder_path=PREL_FILE_PATH):
    """Bloomberg PREL 워크시트(.xlsx)를 읽어 최근 발행 데이터 반환"""
    import glob

    # 폴더에서 PREL 파일 찾기
    search_patterns = [
        os.path.join(folder_path, "*PREL*.xlsx"),
        os.path.join(folder_path, "*Preliminary*.xlsx"),
        os.path.join(folder_path, "*prel*.xlsx"),
    ]

    prel_file = None
    for pattern in search_patterns:
        files = glob.glob(pattern)
        if files:
            # 가장 최근 수정된 파일 선택
            prel_file = max(files, key=os.path.getmtime)
            break

    if not prel_file:
        print(f"  ⚠️  PREL 파일을 찾을 수 없습니다: {folder_path}")
        return []

    print(f"  📄 PREL 파일: {os.path.basename(prel_file)}")

    try:
        import openpyxl
        wb = openpyxl.load_workbook(prel_file, data_only=True, read_only=True)
        ws = wb.active

        # 컬럼 매핑 (1-indexed)
        COL = {
            "date": 1, "cor": 2, "ighy": 3, "ticker": 4, "issuer": 5,
            "maturity": 6, "sp": 7, "ipt": 8, "final_guidance": 9,
            "nic": 10, "amt": 11, "guidance": 12, "latest_level": 13,
            "ccy": 14, "bookrunner": 15, "registration": 16,
            "status": 20, "coupon": 21, "tenor": 24,
            "rank": 29, "updated": 36,
        }

        issues = []
        seen = set()  # 중복 제거
        for r in range(3, ws.max_row + 1):
            ticker = ws.cell(r, COL["ticker"]).value
            if not ticker or ticker == "Most Recent Preliminary Bonds":
                continue

            issuer = ws.cell(r, COL["issuer"]).value
            ighy = ws.cell(r, COL["ighy"]).value
            status = ws.cell(r, COL["status"]).value

            # 중복 키 (ticker + registration)
            reg = ws.cell(r, COL["registration"]).value or ""
            dedup_key = f"{ticker}_{reg}"
            if dedup_key in seen:
                continue
            seen.add(dedup_key)

            def safe_str(v):
                if v is None:
                    return None
                return str(v).strip() if str(v).strip() not in ("None", "nan", "") else None

            def safe_float(v):
                if v is None:
                    return None
                try:
                    return round(float(v), 3)
                except:
                    return None

            issues.append({
                "date": safe_str(ws.cell(r, COL["date"]).value),
                "cor": safe_str(ws.cell(r, COL["cor"]).value),
                "ighy": safe_str(ighy),
                "ticker": safe_str(ticker),
                "issuer": safe_str(issuer),
                "sp": safe_str(ws.cell(r, COL["sp"]).value),
                "ipt": safe_str(ws.cell(r, COL["ipt"]).value),
                "fg": safe_str(ws.cell(r, COL["final_guidance"]).value),
                "nic": safe_float(ws.cell(r, COL["nic"]).value),
                "amt": safe_str(ws.cell(r, COL["amt"]).value),
                "ccy": safe_str(ws.cell(r, COL["ccy"]).value),
                "coupon": safe_float(ws.cell(r, COL["coupon"]).value),
                "tenor": safe_str(ws.cell(r, COL["tenor"]).value),
                "status": safe_str(status),
                "rank": safe_str(ws.cell(r, COL["rank"]).value),
                "reg": safe_str(reg),
            })

        wb.close()
        print(f"  → {len(issues)}건 발행 로드 (IG: {sum(1 for i in issues if i.get('ighy')=='IG')} / HY: {sum(1 for i in issues if i.get('ighy')=='HY')})")
        return issues

    except Exception as e:
        print(f"  ⚠️  PREL 읽기 오류: {e}")
        return []


def get_issues_history_path():
    """누적 발행 데이터 파일 경로"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(script_dir, "new_issues_history.json")


def load_issues_history():
    """기존 누적 발행 데이터 로드"""
    path = get_issues_history_path()
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            pass
    return []


def save_issues_history(issues):
    """누적 발행 데이터 저장"""
    path = get_issues_history_path()
    with open(path, "w", encoding="utf-8") as f:
        json.dump(issues, f, ensure_ascii=False, indent=2)


def merge_new_issues(new_issues):
    """
    오늘 PREL 데이터를 기존 누적 데이터에 병합.
    - ticker 기준으로 중복 제거 (새 데이터가 기존 데이터를 업데이트)
    - 30일 이내 데이터만 유지
    """
    history = load_issues_history()

    # 오늘 날짜 태깅 (PREL의 date는 "05/19" 같은 형식이라 연도가 없음)
    today_str = datetime.today().strftime("%Y-%m-%d")
    year = datetime.today().strftime("%Y")

    for issue in new_issues:
        # full_date가 없으면 추가 (연도 포함)
        if not issue.get("full_date"):
            d = issue.get("date", "")
            if d and "/" in d and len(d) <= 5:
                issue["full_date"] = f"{year}/{d}"
            else:
                issue["full_date"] = today_str
        # 수집 날짜 기록
        if not issue.get("collected"):
            issue["collected"] = today_str

    # 기존 데이터 dict (ticker를 키로)
    existing = {}
    for issue in history:
        key = issue.get("ticker", "")
        if key:
            existing[key] = issue

    # 새 데이터로 업데이트 (같은 ticker면 새 데이터가 우선 — status 등 업데이트 반영)
    for issue in new_issues:
        key = issue.get("ticker", "")
        if key:
            existing[key] = issue

    # 30일 이내만 유지
    cutoff = (datetime.today() - timedelta(days=30)).strftime("%Y-%m-%d")
    merged = []
    for issue in existing.values():
        collected = issue.get("collected", "")
        if collected >= cutoff:
            merged.append(issue)

    # 날짜 역순 정렬
    merged.sort(key=lambda x: x.get("full_date", "") or x.get("collected", ""), reverse=True)

    # 저장
    save_issues_history(merged)
    print(f"  📦 누적 발행: {len(merged)}건 (신규 {len(new_issues)}건 병합, 30일 이내)")

    return merged


# ══════════════════════════════════════════════
# 2. narwhals 헬퍼
# ══════════════════════════════════════════════
def _nw(r):
    try:
        import narwhals as nw
        native = nw.to_native(r)
        if hasattr(native, "to_pandas"):
            return native.to_pandas()
    except:
        pass
    if hasattr(r, "to_pandas"):
        return r.to_pandas()
    return pd.DataFrame(r)


# ══════════════════════════════════════════════
# 3. Bloomberg 데이터 로드
# ══════════════════════════════════════════════
def load_bloomberg_data():
    """Bloomberg에서 모든 데이터를 가져와 JSON 직렬화 가능한 dict로 반환"""
    from xbbg import blp

    result = {}
    print("=" * 60)
    print("  Bloomberg 데이터 로드 시작")
    print("=" * 60)

    # ─── 보유 채권 스냅샷 ───
    print("\n[1/7] 보유 채권 스냅샷 로드 중...")
    tickers = [f"{i} Corp" for i in BOND_ISINS]
    field_values = list(FIELDS.values())
    field_keys = list(FIELDS.keys())

    raw = blp.bdp(tickers, field_values)
    pdf = _nw(raw)
    pdf = pdf.drop_duplicates(subset=["ticker", "field"], keep="last")
    snap = pdf.pivot(index="ticker", columns="field", values="value")
    snap.columns.name = None
    snap = snap.rename(columns=dict(zip(field_values, field_keys)))

    for col in ["oas","spread","ytm","duration","price","coupon","amt_out","zspread"]:
        if col in snap.columns:
            snap[col] = pd.to_numeric(snap[col], errors="coerce")

    if "sector" in snap.columns and "industry" in snap.columns:
        snap["sector"] = snap["sector"].fillna(snap["industry"])
        snap.loc[snap["sector"].astype(str).isin(["nan","None",""]), "sector"] = snap["industry"]
    elif "industry" in snap.columns:
        snap["sector"] = snap["industry"]

    def make_label(r):
        try:
            for col in ["name", "SECURITY_DES"]:
                val = str(r.get(col, "")).strip()
                if val and val not in ("nan", "None", ""):
                    return val
            issuer = str(r.get("issuer", "")).strip()
            coupon = r.get("coupon")
            mat    = r.get("maturity")
            if issuer and issuer not in ("nan","None","") and pd.notna(coupon) and pd.notna(mat):
                return f"{issuer} {float(coupon):.2f} {pd.to_datetime(mat).strftime('%m/%y')}"
            elif issuer and issuer not in ("nan","None",""):
                return issuer
            return r.name.replace(" Corp", "")
        except:
            return r.name.replace(" Corp", "")

    snap["label"] = snap.apply(make_label, axis=1)
    snap["is_owned"] = True

    # 지역 분류
    # CNTRY_OF_RISK 가져오기 시도
    try:
        raw_cor = blp.bdp(tickers, ["CNTRY_OF_RISK"])
        cor_pdf = _nw(raw_cor)
        cor_pdf = cor_pdf.drop_duplicates(subset=["ticker","field"], keep="last")
        cor_map = dict(zip(cor_pdf["ticker"], cor_pdf["value"]))
    except:
        cor_map = {}

    KP_ISSUERS = ["KOOKMIN","KB ","MIRAE","SK HYNIX","SK TELECOM","SK ","SAMSUNG","POSCO","HYUNDAI",
                  "LG CHEM","LG ENERGY","KDB","KEXIM","KOREA","HANWHA","SHINHAN","WOORI","HANA ",
                  "INDUSTRIAL BANK OF KOREA","KOREA ELECTRIC","KEPCO","KOGAS","NPS","KT CORP",
                  "LOTTE","CELLTRION","HYOSUNG","DOOSAN","CJ ","GS ENERGY","KUMHO","DAEWOO"]

    def classify_region(ticker):
        # 1) CNTRY_OF_RISK 우선
        cor = str(cor_map.get(ticker, "")).strip().upper()
        if cor == "KR":
            return "아시아/기타"
        if cor in ("US",):
            return "미국"
        if cor in ("GB","FR","DE","NL","CH","IT","ES","BE","AT","IE","LU","FI","NO","SE","DK"):
            return "유럽"
        if cor in ("JP","CN","HK","SG","AU","TW","IN","ID","TH","MY","PH","NZ"):
            return "아시아/기타"
        if cor and cor not in ("", "NONE", "NAN"):
            return "기타"

        # 2) 발행자 이름 기반 (KP 채권)
        issuer_name = str(snap.loc[ticker, "issuer"] if ticker in snap.index else "").upper()
        if any(kp in issuer_name for kp in KP_ISSUERS):
            return "아시아/기타"

        # 3) ISIN fallback
        isin = str(ticker).replace(" Corp","").strip()
        if isin.startswith("USY") or isin.startswith("USG") or isin.startswith("USH") or isin.startswith("USJ"):
            return "아시아/기타"
        if isin.startswith("US"):
            return "미국"
        if any(isin.startswith(p) for p in ["XS","FR","DE","GB"]):
            return "유럽"
        if any(isin.startswith(p) for p in ["KR","JP","SG","HK"]):
            return "아시아/기타"
        return "기타"

    snap["region"] = snap.index.map(classify_region)

    # 만기 구간
    def maturity_bucket(mat):
        try:
            if pd.isna(mat): return None
            years = (pd.to_datetime(mat) - pd.Timestamp.today()).days / 365.25
            if years < 1: return "1년 이하"
            elif years < 3: return "1-3년"
            elif years < 5: return "3-5년"
            elif years < 7: return "5-7년"
            elif years < 10: return "7-10년"
            elif years < 20: return "10-20년"
            else: return "20년 이상"
        except:
            return None

    snap["matBucket"] = snap["maturity"].apply(maturity_bucket) if "maturity" in snap.columns else None
    print(f"  → {len(snap)}개 종목 로드 완료")

    # ─── 벤치마크 → G-spread 시계열 ───
    print("\n[2/7] 벤치마크 매핑 + G-spread 시계열...")
    try:
        raw_bm = blp.bdp(tickers, ["YAS_BENCHMARK_BOND_ISIN"])
        bm_pdf = _nw(raw_bm)
        bm_pdf = bm_pdf.drop_duplicates(subset=["ticker","field"], keep="last")
        bm_snap = bm_pdf.pivot(index="ticker", columns="field", values="value")

        bm_isins = bm_snap.get("YAS_BENCHMARK_BOND_ISIN", pd.Series()).dropna().unique().tolist()
        bm_isins = [i for i in bm_isins if i]
        if bm_isins:
            raw_bm_mat = blp.bdp([f"{i} Corp" for i in bm_isins], "MATURITY")
            bm_mat_pdf = _nw(raw_bm_mat)
            bm_mat_pdf = bm_mat_pdf.drop_duplicates(subset=["ticker","field"], keep="last")
            bm_mat_map = dict(zip(bm_mat_pdf["ticker"], bm_mat_pdf["value"]))
        else:
            bm_mat_map = {}

        def years_to_generic(years):
            if years is None or pd.isna(years): return "GT10 Govt"
            generic = [(2,"GT2 Govt"),(3,"GT3 Govt"),(5,"GT5 Govt"),
                       (7,"GT7 Govt"),(10,"GT10 Govt"),(20,"GT20 Govt"),(30,"GT30 Govt")]
            return min(generic, key=lambda x: abs(x[0]-years))[1]

        bm_map = {}
        for t in tickers:
            bm_isin = bm_snap.loc[t, "YAS_BENCHMARK_BOND_ISIN"] if t in bm_snap.index else None
            if pd.notna(bm_isin) and bm_isin:
                bm_full = f"{bm_isin} Corp"
                mat = bm_mat_map.get(bm_full)
                if pd.notna(mat) and mat:
                    years = (pd.to_datetime(mat) - pd.Timestamp.today()).days / 365.25
                    bm_map[t] = years_to_generic(years)
                    continue
            if t in snap.index and pd.notna(snap.loc[t].get("duration")):
                bm_map[t] = years_to_generic(float(snap.loc[t]["duration"]))
            else:
                bm_map[t] = "GT10 Govt"

        generic_tickers = sorted(set(bm_map.values()))
    except Exception as e:
        print(f"  벤치마크 조회 오류: {e}")
        bm_map = {t: "GT10 Govt" for t in tickers}
        generic_tickers = ["GT10 Govt"]

    # YTM 시계열 + Generic UST → G-spread 계산
    try:
        raw_h = blp.bdh(tickers, "YLD_YTM_MID",
                        START_DATE.strftime("%Y%m%d"), END_DATE.strftime("%Y%m%d"))
        hpdf = _nw(raw_h)
        bond_ytm_ts = hpdf.pivot(index="date", columns="ticker", values="value")
        bond_ytm_ts.index = pd.to_datetime(bond_ytm_ts.index)
        bond_ytm_ts = bond_ytm_ts.apply(pd.to_numeric, errors="coerce")

        raw_gen = blp.bdh(generic_tickers, "YLD_YTM_MID",
                          START_DATE.strftime("%Y%m%d"), END_DATE.strftime("%Y%m%d"))
        gen_pdf = _nw(raw_gen)
        gen_ytm = gen_pdf.pivot(index="date", columns="ticker", values="value")
        gen_ytm.index = pd.to_datetime(gen_ytm.index)
        gen_ytm = gen_ytm.apply(pd.to_numeric, errors="coerce")

        hist = pd.DataFrame(index=bond_ytm_ts.index)
        label_map = dict(zip(snap.index, snap["label"]))
        for t in tickers:
            gen_t = bm_map.get(t, "GT10 Govt")
            label = label_map.get(t, t)
            if t in bond_ytm_ts.columns and gen_t in gen_ytm.columns:
                hist[label] = (bond_ytm_ts[t] - gen_ytm[gen_t]) * 100  # bps
        hist = hist.dropna(how="all")
        print(f"  → G-spread 시계열: {len(hist)}일, {hist.notna().any().sum()}개 종목")
    except Exception as e:
        print(f"  BDH 오류: {e}")
        hist = pd.DataFrame(index=pd.bdate_range(START_DATE, END_DATE))

    # spread 변동 계산 + 과거 YTM
    for idx, row in snap.iterrows():
        lbl = row["label"]
        if lbl in hist.columns and len(hist) > 0:
            s = hist[lbl].dropna()
            n = len(s)
            snap.at[idx, "chg1d"] = float(s.iloc[-1] - s.iloc[-2]) if n > 1 else None
            snap.at[idx, "chg1w"] = float(s.iloc[-1] - s.iloc[-6]) if n > 5 else None
            snap.at[idx, "chg1m"] = float(s.iloc[-1] - s.iloc[-22]) if n > 21 else None
            snap.at[idx, "chg3m"] = float(s.iloc[-1] - s.iloc[0]) if n > 1 else None
        else:
            snap.at[idx, "chg1d"] = None
            snap.at[idx, "chg1w"] = None
            snap.at[idx, "chg1m"] = None
            snap.at[idx, "chg3m"] = None
        # 1M/3M 전 YTM
        try:
            if idx in bond_ytm_ts.columns:
                ytm_s = bond_ytm_ts[idx].dropna()
                snap.at[idx, "ytm_1m_ago"] = float(ytm_s.iloc[-22]) if len(ytm_s) > 21 else None
                snap.at[idx, "ytm_3m_ago"] = float(ytm_s.iloc[0]) if len(ytm_s) > 1 else None
            else:
                snap.at[idx, "ytm_1m_ago"] = None
                snap.at[idx, "ytm_3m_ago"] = None
        except:
            snap.at[idx, "ytm_1m_ago"] = None
            snap.at[idx, "ytm_3m_ago"] = None

    # ─── UST 수익률 곡선 ───
    print("\n[3/7] UST 수익률 곡선...")
    try:
        ust_t = list(UST_TICKERS.values())
        raw_u = blp.bdp(ust_t, "YLD_YTM_MID")
        updf = _nw(raw_u)
        ust_now_map = dict(zip(updf["ticker"], pd.to_numeric(updf["value"], errors="coerce")))

        raw_u3 = blp.bdh(ust_t, "YLD_YTM_MID",
                         (END_DATE - timedelta(days=92)).strftime("%Y%m%d"),
                         (END_DATE - timedelta(days=88)).strftime("%Y%m%d"))
        u3pdf = _nw(raw_u3)
        ust3m_map = (u3pdf.groupby("ticker")["value"].last()
                         .apply(pd.to_numeric, errors="coerce").to_dict())

        ust_now = [float(ust_now_map.get(t, float("nan"))) for t in ust_t]
        ust_3m_ago = [float(ust3m_map.get(t, float("nan"))) for t in ust_t]
        print(f"  → UST 곡선 로드 완료")
    except Exception as e:
        print(f"  UST 곡선 오류: {e}")
        ust_now = [5.31,5.28,5.12,4.85,4.62,4.52,4.41,4.38,4.32,4.51,4.55]
        ust_3m_ago = [5.05,5.10,4.95,4.68,4.44,4.33,4.25,4.22,4.18,4.37,4.43]

    # ─── 발행자 곡선용 추가 채권 ───
    print("\n[4/7] 발행자 곡선 추가 채권...")
    extra_tickers = [f"{i} Corp" for i in ISSUER_CURVE_ISINS if i not in BOND_ISINS]
    extra_snap = pd.DataFrame()
    extra_ytm_ts = pd.DataFrame()
    if extra_tickers:
        try:
            raw_ex = blp.bdp(extra_tickers, field_values)
            ex_pdf = _nw(raw_ex)
            ex_pdf = ex_pdf.drop_duplicates(subset=["ticker","field"], keep="last")
            extra_snap = ex_pdf.pivot(index="ticker", columns="field", values="value")
            extra_snap.columns.name = None
            extra_snap = extra_snap.rename(columns=dict(zip(field_values, field_keys)))
            for col in ["oas","spread","ytm","duration","price","coupon"]:
                if col in extra_snap.columns:
                    extra_snap[col] = pd.to_numeric(extra_snap[col], errors="coerce")
            extra_snap["label"] = extra_snap.apply(make_label, axis=1)
            extra_snap["is_owned"] = False

            # 지역 분류 (발행자 이름 기반 포함)
            def classify_extra_region(ticker):
                issuer_name = str(extra_snap.loc[ticker, "issuer"] if ticker in extra_snap.index else "").upper()
                if any(kp in issuer_name for kp in KP_ISSUERS):
                    return "아시아/기타"
                return classify_region(ticker)
            extra_snap["region"] = extra_snap.index.map(classify_extra_region)
            extra_snap["matBucket"] = extra_snap["maturity"].apply(maturity_bucket) if "maturity" in extra_snap.columns else None

            # YTM 시계열
            raw_exh = blp.bdh(extra_tickers, "YLD_YTM_MID",
                              START_DATE.strftime("%Y%m%d"), END_DATE.strftime("%Y%m%d"))
            exh_pdf = _nw(raw_exh)
            extra_ytm_ts = exh_pdf.pivot(index="date", columns="ticker", values="value")
            extra_ytm_ts.index = pd.to_datetime(extra_ytm_ts.index)
            extra_ytm_ts = extra_ytm_ts.apply(pd.to_numeric, errors="coerce")
            ex_label_map = dict(zip(extra_snap.index, extra_snap["label"]))
            extra_ytm_ts.columns = [ex_label_map.get(c, c) for c in extra_ytm_ts.columns]

            # 추가 채권 과거 YTM
            for idx, row in extra_snap.iterrows():
                try:
                    if idx in extra_ytm_ts.columns or ex_label_map.get(idx) in extra_ytm_ts.columns:
                        col = ex_label_map.get(idx, idx)
                        if col in extra_ytm_ts.columns:
                            ytm_s = extra_ytm_ts[col].dropna()
                            extra_snap.at[idx, "ytm_1m_ago"] = float(ytm_s.iloc[-22]) if len(ytm_s) > 21 else None
                            extra_snap.at[idx, "ytm_3m_ago"] = float(ytm_s.iloc[0]) if len(ytm_s) > 1 else None
                            continue
                    extra_snap.at[idx, "ytm_1m_ago"] = None
                    extra_snap.at[idx, "ytm_3m_ago"] = None
                except:
                    extra_snap.at[idx, "ytm_1m_ago"] = None
                    extra_snap.at[idx, "ytm_3m_ago"] = None

            # 중복 제거
            extra_snap = extra_snap[~extra_snap.index.isin(snap.index)]
            print(f"  → 추가 {len(extra_snap)}개 채권 로드 완료")
        except Exception as e:
            print(f"  추가 채권 오류: {e}")

    # ─── 시장 지표 ───
    print("\n[5/7] 시장 지표 (VIX, MOVE, CDX 등)...")
    mkt_ts = {}
    mkt_current = {}
    mkt_changes = {}
    for tk, (name, field) in MARKET_TICKERS.items():
        try:
            r = blp.bdh([tk], field, MKT_START.strftime("%Y%m%d"), END_DATE.strftime("%Y%m%d"))
            mpdf = _nw(r)
            if len(mpdf) > 0:
                s = mpdf.set_index("date")["value"]
                s = pd.to_numeric(s, errors="coerce").dropna().sort_index()
                s.index = pd.to_datetime(s.index)
                if len(s) > 0:
                    mkt_ts[name] = {d.strftime("%Y-%m-%d"): float(v) for d, v in s.items()}
                    mkt_current[name] = float(s.iloc[-1])
                    chg = {}
                    for lbl, n in [("1d",1),("1w",5),("1m",21)]:
                        chg[lbl] = float(s.iloc[-1] - s.iloc[-(n+1)]) if len(s) > n else None
                    mkt_changes[name] = chg
                    continue
        except Exception as e:
            print(f"  {name} 오류: {e}")
        mkt_current[name] = None
        mkt_changes[name] = {"1d": None, "1w": None, "1m": None}
    print(f"  → {len(mkt_current)}개 지표 로드")

    # ─── 거래 데이터 - TRACE (최근 7일) ───
    print("\n[6/8] TRACE 거래 데이터...")
    trades = []
    for tk in tickers:
        try:
            # TRACE 거래 내역: TRAC <GO> 에서 제공하는 필드
            r = blp.bdh([tk],
                        ["TRACE_DLR_LAST_PX", "TRACE_DLR_LAST_YLD", "TRACE_DLR_LAST_SPREAD",
                         "TRACE_DLR_LAST_VOL", "TRACE_DLR_RPT_PARTY"],
                        (datetime.today() - timedelta(days=7)).strftime("%Y%m%d"),
                        datetime.today().strftime("%Y%m%d"))
            tpdf = _nw(r)
            if len(tpdf) == 0:
                continue
            df = tpdf.pivot_table(index="date", columns="field", values="value", aggfunc="last").reset_index()
            lbl = snap.loc[tk, "label"] if tk in snap.index else tk
            yas_spread = pd.to_numeric(snap.loc[tk, "spread"], errors="coerce") if tk in snap.index else None

            for _, row in df.iterrows():
                px = pd.to_numeric(row.get("TRACE_DLR_LAST_PX"), errors="coerce")
                ytm_val = pd.to_numeric(row.get("TRACE_DLR_LAST_YLD"), errors="coerce")
                trd_spread = pd.to_numeric(row.get("TRACE_DLR_LAST_SPREAD"), errors="coerce")
                vol = pd.to_numeric(row.get("TRACE_DLR_LAST_VOL"), errors="coerce")
                source = row.get("TRACE_DLR_RPT_PARTY")

                # 거래 스프레드 vs 전일 YAS spread
                spread_chg = None
                if pd.notna(trd_spread) and pd.notna(yas_spread):
                    spread_chg = round(float(trd_spread - yas_spread), 1)

                trades.append({
                    "date": pd.to_datetime(row["date"]).strftime("%Y-%m-%d"),
                    "bond": lbl,
                    "price": round(float(px), 3) if pd.notna(px) else None,
                    "ytm": round(float(ytm_val), 3) if pd.notna(ytm_val) else None,
                    "trd_spread": round(float(trd_spread), 1) if pd.notna(trd_spread) else None,
                    "yas_spread": round(float(yas_spread), 1) if pd.notna(yas_spread) else None,
                    "spread_chg": spread_chg,
                    "volume": int(vol) if pd.notna(vol) and vol > 0 else None,
                    "source": str(source).strip() if pd.notna(source) and str(source).strip() not in ("nan","None","") else None,
                })
        except:
            pass
    trades.sort(key=lambda x: x["date"], reverse=True)
    print(f"  → {len(trades)}건 TRACE 거래 로드")

    # ─── ETF 스프레드 시계열 ───
    print("\n[7/8] ETF 스프레드 데이터...")
    etf_ts = {}
    etf_current = {}
    etf_changes = {}
    for tk, name in ETF_TICKERS.items():
        try:
            r = blp.bdh([tk], "YAS_YLD_SPREAD", MKT_START.strftime("%Y%m%d"), END_DATE.strftime("%Y%m%d"))
            epdf = _nw(r)
            if len(epdf) > 0:
                s = epdf.set_index("date")["value"]
                s = pd.to_numeric(s, errors="coerce").dropna().sort_index()
                s.index = pd.to_datetime(s.index)
                if len(s) > 0:
                    etf_ts[name] = {d.strftime("%Y-%m-%d"): float(v) for d, v in s.items()}
                    etf_current[name] = float(s.iloc[-1])
                    chg = {}
                    for lbl, n in [("1d",1),("1w",5),("1m",21)]:
                        chg[lbl] = float(s.iloc[-1] - s.iloc[-(n+1)]) if len(s) > n else None
                    chg["3m"] = float(s.iloc[-1] - s.iloc[-64]) if len(s) > 63 else None
                    chg["6m"] = float(s.iloc[-1] - s.iloc[0]) if len(s) > 1 else None
                    etf_changes[name] = chg
                    continue
        except Exception as e:
            print(f"  {name} 오류: {e}")
        etf_current[name] = None
        etf_changes[name] = {"1d": None, "1w": None, "1m": None}
    print(f"  → {len(etf_current)}개 ETF 로드")

    # ─── BDC 데이터 ───
    print("\n[8/9] BDC 주가 / NAV 할인율 로드...")
    bdc_current = {}
    bdc_changes = {}
    bdc_ts = {}
    bdc_tickers_list = list(BDC_TICKERS.keys())
    try:
        ref_raw = blp.bdp(bdc_tickers_list,
                          ["PX_LAST", "PX_TO_BOOK_RATIO", "NET_ASSET_VAL",
                           "CHG_PCT_1D", "CHG_PCT_5D", "CHG_PCT_1M"])
        ref_pdf = _nw(ref_raw)
        ref_pdf = ref_pdf.drop_duplicates(subset=["ticker","field"], keep="last")
        ref_pivot = ref_pdf.pivot(index="ticker", columns="field", values="value")
        ref_pivot.columns.name = None
        for col in ref_pivot.columns:
            ref_pivot[col] = pd.to_numeric(ref_pivot[col], errors="coerce")
        for tk, meta in BDC_TICKERS.items():
            nm = meta["short"]
            bdc_current[nm] = {}
            bdc_changes[nm] = {}
            if tk in ref_pivot.index:
                row = ref_pivot.loc[tk]
                for fld, key in [("PX_LAST","price"),("PX_TO_BOOK_RATIO","pb"),("NET_ASSET_VAL","nav")]:
                    v = row.get(fld)
                    bdc_current[nm][key] = round(float(v), 3) if pd.notna(v) else None
                for fld, key in [("CHG_PCT_1D","1d"),("CHG_PCT_5D","1w"),("CHG_PCT_1M","1m")]:
                    v = row.get(fld)
                    bdc_changes[nm][key] = round(float(v), 2) if pd.notna(v) else None
        print(f"  -> BDC {len([k for k,v in bdc_current.items() if v.get(chr(39)+'price'+chr(39))])}개 로드")
    except Exception as e:
        print(f"  BDC BDP 오류: {e}")

    try:
        ts_raw = blp.bdh(bdc_tickers_list, "PX_LAST",
                         MKT_START.strftime("%Y%m%d"), END_DATE.strftime("%Y%m%d"))
        ts_pdf = _nw(ts_raw)
        ts_pdf = ts_pdf.drop_duplicates(subset=["ticker","date"], keep="last")
        for tk, meta in BDC_TICKERS.items():
            nm = meta["short"]
            sub = ts_pdf[ts_pdf["ticker"] == tk].copy()
            if len(sub) > 0:
                s = pd.to_numeric(sub.set_index("date")["value"], errors="coerce").dropna()
                s.index = pd.to_datetime(s.index)
                s = s.sort_index()
                if len(s) > 5:
                    bdc_ts[nm] = {d.strftime("%Y-%m-%d"): round(float(v), 3) for d, v in s.items()}
        print(f"  -> BDC 시계열 {len(bdc_ts)}개 로드")
    except Exception as e:
        print(f"  BDC BDH 오류: {e}")

    # ─── JSON 직렬화 ───
    print("\n[9/9] JSON 변환 중...")

    # PREL 워크시트 로드 + 누적
    print("\n  📋 PREL 최근 발행 데이터 로드...")
    today_issues = load_prel_data()
    new_issues = merge_new_issues(today_issues) if today_issues else load_issues_history()

    # 보유 + 추가 합치기
    combined = pd.concat([snap, extra_snap]) if not extra_snap.empty else snap.copy()

    def bonds_to_json(df):
        records = []
        for idx, row in df.iterrows():
            rec = {}
            # ISIN from ticker index (e.g. "US500630ED65 Corp" → "US500630ED65")
            rec["isin"] = str(idx).replace(" Corp","").strip() if idx else None
            for c in ["label","issuer","sector","rating","region","matBucket","is_owned"]:
                v = row.get(c)
                rec[c] = str(v) if pd.notna(v) and str(v) not in ("nan","None") else None
            for c in ["spread","ytm","duration","price","coupon","oas","chg1d","chg1w","chg1m","chg3m","ytm_1m_ago","ytm_3m_ago"]:
                v = row.get(c)
                rec[c] = round(float(v), 3) if pd.notna(v) else None
            rec["is_owned"] = bool(row.get("is_owned", True))
            records.append(rec)
        return records

    def ts_to_json(df):
        if df.empty:
            return {"dates": [], "series": {}}
        dates = [d.strftime("%Y-%m-%d") for d in df.index]
        series = {}
        for col in df.columns:
            series[col] = [round(float(v), 2) if pd.notna(v) else None for v in df[col]]
        return {"dates": dates, "series": series}

    result = {
        "generated": datetime.today().strftime("%Y-%m-%d %H:%M"),
        "bondCount": len(snap),
        "bondIsins": BOND_ISINS,
        "bonds": bonds_to_json(snap),
        "allBonds": bonds_to_json(combined),
        "spreadHistory": ts_to_json(hist),
        "ustCurve": {
            "tenors": list(UST_TICKERS.keys()),
            "years": [1/12,.25,.5,1,2,3,5,7,10,20,30],
            "now": [round(v, 3) if not np.isnan(v) else None for v in ust_now],
            "ago3m": [round(v, 3) if not np.isnan(v) else None for v in ust_3m_ago],
        },
        "marketTimeseries": mkt_ts,
        "marketCurrent": {k: round(v, 2) if v else None for k, v in mkt_current.items()},
        "marketChanges": mkt_changes,
        "etfTimeseries": etf_ts,
        "etfCurrent": {k: round(v, 2) if v else None for k, v in etf_current.items()},
        "etfChanges": etf_changes,
        "trades": trades,
        "newIssues": new_issues,
        "bdcCurrent": bdc_current,
        "bdcChanges": bdc_changes,
        "bdcTimeseries": bdc_ts,
        "bdcMeta": {meta["short"]: {"name": meta["name"], "color": meta["color"]} for meta in BDC_TICKERS.values()},
    }

    print(f"\n{'='*60}")
    print(f"  데이터 로드 완료!")
    print(f"  보유 채권: {len(snap)}개")
    print(f"  추가 채권: {len(extra_snap)}개")
    print(f"  시계열:    {len(hist)}일")
    print(f"  시장 지표: {len(mkt_current)}개")
    print(f"  ETF:       {len(etf_current)}개")
    print(f"  최근 발행: {len(new_issues)}건")
    print(f"  거래:      {len(trades)}건")
    print(f"{'='*60}")

    return result


def load_sample_data():
    """Bloomberg 없을 때 샘플 데이터 생성"""
    np.random.seed(42)

    def maturity_bucket_sample(dur):
        if dur < 1: return "1년 이하"
        elif dur < 3: return "1-3년"
        elif dur < 5: return "3-5년"
        elif dur < 7: return "5-7년"
        elif dur < 10: return "7-10년"
        elif dur < 20: return "10-20년"
        else: return "20년 이상"

    ISSUERS = ['HSBC','JPMorgan','Goldman Sachs','Morgan Stanley','BNP Paribas',
               'Samsung','SK Hynix','Hyundai','POSCO','LG Chem','Toyota','Mizuho',
               'Barclays','Deutsche Bank','Credit Agricole','Societe Generale',
               'ING','UBS','Standard Chartered','ANZ']
    SECTORS = ['Financials','Technology','Energy','Utilities','Materials','Consumer','Industrials','Communications']
    RATINGS = ['AA+','AA','AA-','A+','A','A-','BBB+','BBB','BBB-']
    REGIONS = ['미국','유럽','아시아/기타']
    MAT_BUCKETS = ['1-3년','3-5년','5-7년','7-10년','10-20년']

    KP_SAMPLE = ['Samsung','SK Hynix','Hyundai','POSCO','LG Chem']
    EU_SAMPLE = ['HSBC','BNP Paribas','Barclays','Deutsche Bank','Credit Agricole','Societe Generale','ING','UBS','Standard Chartered','ANZ']
    US_SAMPLE = ['JPMorgan','Goldman Sachs','Morgan Stanley']
    def sample_region(issuer):
        if issuer in KP_SAMPLE: return '아시아/기타'
        if issuer in EU_SAMPLE: return '유럽'
        if issuer in US_SAMPLE: return '미국'
        return '기타'

    bonds = []
    for i, isin in enumerate(BOND_ISINS):
        issuer = ISSUERS[i % len(ISSUERS)]
        cpn = round(2.0 + np.random.random() * 4.5, 2)
        mat_y = 25 + np.random.randint(0, 10)
        mat_m = str(1 + np.random.randint(0,12)).zfill(2)
        bonds.append({
            "isin": isin,
            "label": f"{issuer} {cpn:.2f} {mat_m}/{mat_y}",
            "issuer": issuer,
            "sector": np.random.choice(SECTORS),
            "rating": np.random.choice(RATINGS),
            "region": sample_region(issuer),
            "matBucket": np.random.choice(MAT_BUCKETS),
            "spread": int(70 + np.random.random() * 250),
            "duration": round(1.5 + np.random.random() * 8, 1),
            "ytm": None,
            "price": round(88 + np.random.random() * 16, 3),
            "coupon": cpn,
            "oas": None,
            "is_owned": True,
        })
    for b in bonds:
        b["ytm"] = round(4.3 + b["spread"] / 100, 2)
        b["oas"] = b["spread"]
        b["ytm_1m_ago"] = round(b["ytm"] + (np.random.random() - 0.5) * 0.3, 2)
        b["ytm_3m_ago"] = round(b["ytm"] + (np.random.random() - 0.5) * 0.5, 2)

    # 추가 채권 (발행자 곡선용) — 실제 매핑
    all_bonds = list(bonds)

    # 추가 종목 실제 데이터 매핑
    EXTRA_BOND_DATA = {
        "ZO621460": {"issuer":"BNP Paribas","label":"BNP 1.904 09/30/28","coupon":1.904,"maturity":"2028-09-30","sector":"Financials"},
        "ZK896516": {"issuer":"BNP Paribas","label":"BNP 5.335 06/12/29","coupon":5.335,"maturity":"2029-06-12","sector":"Financials"},
        "ZF990477": {"issuer":"BNP Paribas","label":"BNP 5.497 05/20/30","coupon":5.497,"maturity":"2030-05-20","sector":"Financials"},
        "YO128804": {"issuer":"BNP Paribas","label":"BNP 5.085 05/09/31","coupon":5.085,"maturity":"2031-05-09","sector":"Financials"},
        "BP054719": {"issuer":"BNP Paribas","label":"BNP 2.871 04/19/32","coupon":2.871,"maturity":"2032-04-19","sector":"Financials"},
        "YS422819": {"issuer":"BNP Paribas","label":"BNP 5.786 01/13/33","coupon":5.786,"maturity":"2033-01-13","sector":"Financials"},
        "DC822725": {"issuer":"BNP Paribas","label":"BNP 4.916 01/15/34","coupon":4.916,"maturity":"2034-01-15","sector":"Financials"},
        "ZF990475": {"issuer":"BNP Paribas","label":"BNP 5.738 02/20/35","coupon":5.738,"maturity":"2035-02-20","sector":"Financials"},
        "YV600775": {"issuer":"Credit Agricole","label":"ACAFP 4.631 09/11/28","coupon":4.631,"maturity":"2028-09-11","sector":"Financials"},
        "ZH098092": {"issuer":"Credit Agricole","label":"ACAFP 6.316 10/03/29","coupon":6.316,"maturity":"2029-10-03","sector":"Financials"},
        "ZF097197": {"issuer":"Credit Agricole","label":"ACAFP 5.335 01/10/30","coupon":5.335,"maturity":"2030-01-10","sector":"Financials"},
        "YO680597": {"issuer":"Credit Agricole","label":"ACAFP 5.222 05/27/31","coupon":5.222,"maturity":"2031-05-27","sector":"Financials"},
        "DC696922": {"issuer":"Credit Agricole","label":"ACAFP 4.656 01/12/32","coupon":4.656,"maturity":"2032-01-12","sector":"Financials"},
        "ZJ311931": {"issuer":"Credit Agricole","label":"ACAFP 5.514 07/05/33","coupon":5.514,"maturity":"2033-07-05","sector":"Financials"},
        "YS344475": {"issuer":"Credit Agricole","label":"ACAFP 5.862 01/09/36","coupon":5.862,"maturity":"2036-01-09","sector":"Financials"},
        "DC696917": {"issuer":"Credit Agricole","label":"ACAFP 5.261 01/12/37","coupon":5.261,"maturity":"2037-01-12","sector":"Financials"},
        "ZI782661": {"issuer":"Barclays","label":"BACR 6.496 09/13/27","coupon":6.496,"maturity":"2027-09-13","sector":"Financials"},
        "ZD475756": {"issuer":"Barclays","label":"BACR 5.674 03/12/28","coupon":5.674,"maturity":"2028-03-12","sector":"Financials"},
        "AS592157": {"issuer":"Barclays","label":"BACR 4.972 05/16/29","coupon":4.972,"maturity":"2029-05-16","sector":"Financials"},
        "DF616766": {"issuer":"Barclays","label":"BACR 4.219 05/24/30","coupon":4.219,"maturity":"2030-05-24","sector":"Financials"},
        "YV564962": {"issuer":"Barclays","label":"BACR 4.942 09/10/30","coupon":4.942,"maturity":"2030-09-10","sector":"Financials"},
        "BK129468": {"issuer":"Barclays","label":"BACR 2.645 06/24/31","coupon":2.645,"maturity":"2031-06-24","sector":"Financials"},
        "BO389986": {"issuer":"Barclays","label":"BACR 2.667 03/10/32","coupon":2.667,"maturity":"2032-03-10","sector":"Financials"},
        "ZN017996": {"issuer":"Barclays","label":"BACR 7.437 11/02/33","coupon":7.437,"maturity":"2033-11-02","sector":"Financials"},
        "ZK471433": {"issuer":"Barclays","label":"BACR 6.224 05/09/34","coupon":6.224,"maturity":"2034-05-09","sector":"Financials"},
        "YV564963": {"issuer":"Barclays","label":"BACR 5.335 09/10/35","coupon":5.335,"maturity":"2035-09-10","sector":"Financials"},
        "YR685574": {"issuer":"Barclays","label":"BACR 5.785 02/25/36","coupon":5.785,"maturity":"2036-02-25","sector":"Financials"},
        "BP135680": {"issuer":"Bank of America","label":"BAC 1.734 07/22/27","coupon":1.734,"maturity":"2027-07-22","sector":"Financials"},
        "BU068258": {"issuer":"Bank of America","label":"BAC 2.551 02/04/28","coupon":2.551,"maturity":"2028-02-04","sector":"Financials"},
        "ZN225587": {"issuer":"Bank of America","label":"BAC 6.204 11/10/28","coupon":6.204,"maturity":"2028-11-10","sector":"Financials"},
        "YO270397": {"issuer":"Bank of America","label":"BAC 4.623 05/09/29","coupon":4.623,"maturity":"2029-05-09","sector":"Financials"},
        "DI841976": {"issuer":"Bank of America","label":"BAC 4.477 04/23/30","coupon":4.477,"maturity":"2030-04-23","sector":"Financials"},
        "YS765590": {"issuer":"Bank of America","label":"BAC 5.162 01/24/31","coupon":5.162,"maturity":"2031-01-24","sector":"Financials"},
        "YI718036": {"issuer":"Bank of America","label":"BAC 4.456 02/06/32","coupon":4.456,"maturity":"2032-02-06","sector":"Financials"},
        "DI841979": {"issuer":"Bank of America","label":"BAC 4.695 04/23/32","coupon":4.695,"maturity":"2032-04-23","sector":"Financials"},
        "BU068259": {"issuer":"Bank of America","label":"BAC 2.972 02/04/33","coupon":2.972,"maturity":"2033-02-04","sector":"Financials"},
        "ZK292754": {"issuer":"Bank of America","label":"BAC 5.288 04/25/34","coupon":5.288,"maturity":"2034-04-25","sector":"Financials"},
        "ZF435445": {"issuer":"Bank of America","label":"BAC 5.468 01/23/35","coupon":5.468,"maturity":"2035-01-23","sector":"Financials"},
        "YO270399": {"issuer":"Bank of America","label":"BAC 5.464 05/09/36","coupon":5.464,"maturity":"2036-05-09","sector":"Financials"},
        "YP652772": {"issuer":"Goldman Sachs","label":"GS 4.937 04/23/28","coupon":4.937,"maturity":"2028-04-23","sector":"Financials"},
        "YI082278": {"issuer":"Goldman Sachs","label":"GS 4.148 01/21/29","coupon":4.148,"maturity":"2029-01-21","sector":"Financials"},
        "YJ138048": {"issuer":"Goldman Sachs","label":"GS 4.153 10/21/29","coupon":4.153,"maturity":"2029-10-21","sector":"Financials"},
        "DI424234": {"issuer":"Goldman Sachs","label":"GS 4.594 04/20/30","coupon":4.594,"maturity":"2030-04-20","sector":"Financials"},
        "YP652776": {"issuer":"Goldman Sachs","label":"GS 5.218 04/23/31","coupon":5.218,"maturity":"2031-04-23","sector":"Financials"},
        "YI082280": {"issuer":"Goldman Sachs","label":"GS 4.516 01/21/32","coupon":4.516,"maturity":"2032-01-21","sector":"Financials"},
        "BT631348": {"issuer":"Goldman Sachs","label":"GS 3.102 02/24/33","coupon":3.102,"maturity":"2033-02-24","sector":"Financials"},
        "DI424237": {"issuer":"Goldman Sachs","label":"GS 5.094 04/20/34","coupon":5.094,"maturity":"2034-04-20","sector":"Financials"},
        "YJ138052": {"issuer":"Goldman Sachs","label":"GS 4.939 10/21/36","coupon":4.939,"maturity":"2036-10-21","sector":"Financials"},
        "YI082283": {"issuer":"Goldman Sachs","label":"GS 5.065 01/21/37","coupon":5.065,"maturity":"2037-01-21","sector":"Financials"},
        "BW306958": {"issuer":"UBS","label":"UBS 4.751 05/12/28","coupon":4.751,"maturity":"2028-05-12","sector":"Financials"},
        "ZI966557": {"issuer":"UBS","label":"UBS 6.246 09/22/29","coupon":6.246,"maturity":"2029-09-22","sector":"Financials"},
        "YK266716": {"issuer":"UBS","label":"UBS 4.151 12/23/29","coupon":4.151,"maturity":"2029-12-23","sector":"Financials"},
        "YI794680": {"issuer":"UBS","label":"UBS 4.214 04/10/30","coupon":4.214,"maturity":"2030-04-10","sector":"Financials"},
        "ZB975719": {"issuer":"UBS","label":"UBS 5.617 09/13/30","coupon":5.617,"maturity":"2030-09-13","sector":"Financials"},
        "YK267609": {"issuer":"UBS","label":"UBS 4.398 09/23/31","coupon":4.398,"maturity":"2031-09-23","sector":"Financials"},
        "YI794697": {"issuer":"UBS","label":"UBS 4.588 08/10/32","coupon":4.588,"maturity":"2032-08-10","sector":"Financials"},
        "YJ946688": {"issuer":"UBS","label":"UBS 4.844 11/06/33","coupon":4.844,"maturity":"2033-11-06","sector":"Financials"},
        "ZI966562": {"issuer":"UBS","label":"UBS 6.301 09/22/34","coupon":6.301,"maturity":"2034-09-22","sector":"Financials"},
        "ZF069836": {"issuer":"UBS","label":"UBS 5.699 02/08/35","coupon":5.699,"maturity":"2035-02-08","sector":"Financials"},
        "YO269253": {"issuer":"UBS","label":"UBS 5.58 05/09/36","coupon":5.58,"maturity":"2036-05-09","sector":"Financials"},
        "YI794696": {"issuer":"UBS","label":"UBS 5.199 08/10/37","coupon":5.199,"maturity":"2037-08-10","sector":"Financials"},
        "ZB470534": {"issuer":"Wells Fargo","label":"WFC 5.707 04/22/28","coupon":5.707,"maturity":"2028-04-22","sector":"Financials"},
        "YP651966": {"issuer":"Wells Fargo","label":"WFC 4.97 04/23/29","coupon":4.97,"maturity":"2029-04-23","sector":"Financials"},
        "YL980741": {"issuer":"Wells Fargo","label":"WFC 4.078 09/15/29","coupon":4.078,"maturity":"2029-09-15","sector":"Financials"},
        "YI082274": {"issuer":"Wells Fargo","label":"WFC 4.182 01/23/30","coupon":4.182,"maturity":"2030-01-23","sector":"Financials"},
        "DK712152": {"issuer":"Wells Fargo","label":"WFC 4.844 05/20/32","coupon":4.844,"maturity":"2032-05-20","sector":"Financials"},
        "BX921089": {"issuer":"Wells Fargo","label":"WFC 4.897 07/25/33","coupon":4.897,"maturity":"2033-07-25","sector":"Financials"},
        "ZJ884878": {"issuer":"Wells Fargo","label":"WFC 5.557 07/25/34","coupon":5.557,"maturity":"2034-07-25","sector":"Financials"},
        "ZF374824": {"issuer":"Wells Fargo","label":"WFC 5.499 01/23/35","coupon":5.499,"maturity":"2035-01-23","sector":"Financials"},
        "YP651972": {"issuer":"Wells Fargo","label":"WFC 5.605 04/23/36","coupon":5.605,"maturity":"2036-04-23","sector":"Financials"},
        "YI082276": {"issuer":"Wells Fargo","label":"WFC 4.96 01/23/37","coupon":4.96,"maturity":"2037-01-23","sector":"Financials"},
    }

    extra_isins = [i for i in ISSUER_CURVE_ISINS if i not in BOND_ISINS]
    for i, isin in enumerate(extra_isins):
        edata = EXTRA_BOND_DATA.get(isin)
        if edata:
            # Use real data
            mat_dt = pd.to_datetime(edata["maturity"])
            dur = max(0.5, round((mat_dt - pd.Timestamp.today()).days / 365.25, 1))
            sp = int(60 + dur * 12 + np.random.random() * 30)
            all_bonds.append({
                "isin": isin,
                "label": edata["label"],
                "issuer": edata["issuer"],
                "sector": edata["sector"],
                "rating": np.random.choice(['A+','A','A-','BBB+']),
                "region": {"BNP Paribas":"유럽","Credit Agricole":"유럽","Barclays":"유럽","Bank of America":"미국","Goldman Sachs":"미국","UBS":"유럽","Wells Fargo":"미국"}.get(edata["issuer"],"미국"),
                "matBucket": maturity_bucket_sample(dur),
                "spread": sp,
                "duration": dur,
                "ytm": round(4.3 + sp / 100, 2),
                "price": round(88 + np.random.random() * 16, 3),
                "coupon": edata["coupon"],
                "oas": sp,
                "is_owned": False,
                "ytm_1m_ago": round(4.3 + sp / 100 + (np.random.random() - 0.5) * 0.3, 2),
                "ytm_3m_ago": round(4.3 + sp / 100 + (np.random.random() - 0.5) * 0.5, 2),
            })
        else:
            # Fallback for ISIN-based entries without explicit mapping
            issuer = ISSUERS[i % len(ISSUERS)]
            cpn = round(2.0 + np.random.random() * 4.5, 2)
            mat_y = 25 + np.random.randint(0, 10)
            mat_m = str(1 + np.random.randint(0,12)).zfill(2)
            dur = round(1.5 + np.random.random() * 8, 1)
            sp = int(60 + dur * 12 + np.random.random() * 30)
            ytm_val = round(4.3 + sp / 100, 2)
            all_bonds.append({
                "isin": isin,
                "label": f"{issuer} {cpn:.2f} {mat_m}/{mat_y}",
                "issuer": issuer,
                "sector": np.random.choice(SECTORS),
                "rating": np.random.choice(RATINGS),
                "region": sample_region(issuer),
                "matBucket": np.random.choice(MAT_BUCKETS),
                "spread": sp,
                "duration": dur,
                "ytm": ytm_val,
                "price": round(88 + np.random.random() * 16, 3),
                "coupon": cpn,
                "oas": sp,
                "is_owned": False,
                "ytm_1m_ago": round(ytm_val + (np.random.random() - 0.5) * 0.3, 2),
                "ytm_3m_ago": round(ytm_val + (np.random.random() - 0.5) * 0.5, 2),
            })

    # 스프레드 시계열
    dates = pd.bdate_range(START_DATE, END_DATE)
    date_strs = [d.strftime("%Y-%m-%d") for d in dates]
    series = {}
    for b in bonds:
        base = b["spread"]
        noise = np.cumsum(np.random.randn(len(dates)) * 3)
        vals = np.clip(base + noise - noise[-1], base * 0.5, base * 1.6)
        series[b["label"]] = [round(float(v), 1) for v in vals]

    # 변동 계산
    for b in bonds:
        s = series.get(b["label"], [])
        n = len(s)
        b["chg1d"] = round(s[-1] - s[-2], 1) if n > 1 else None
        b["chg1w"] = round(s[-1] - s[-6], 1) if n > 5 else None
        b["chg1m"] = round(s[-1] - s[-22], 1) if n > 21 else None
        b["chg3m"] = round(s[-1] - s[0], 1) if n > 1 else None

    # 시장 지표
    mkt_dates = pd.bdate_range(MKT_START, END_DATE)
    mkt_date_strs = [d.strftime("%Y-%m-%d") for d in mkt_dates]
    mkt_ts = {}
    mkt_bases = {"US IG OAS": 95, "US HY OAS": 380, "CDX IG 5Y": 55, "CDX HY 5Y": 340, "VIX": 18, "MOVE": 105}
    mkt_vols = {"US IG OAS": 3, "US HY OAS": 12, "CDX IG 5Y": 2, "CDX HY 5Y": 10, "VIX": 2, "MOVE": 4}
    mkt_current = {}
    mkt_changes = {}
    for name, base in mkt_bases.items():
        vals = [base]
        for i in range(1, len(mkt_dates)):
            vals.append(max(base * 0.6, min(base * 1.5, vals[-1] + (np.random.random() - 0.48) * mkt_vols[name])))
        mkt_ts[name] = {mkt_date_strs[i]: round(float(v), 1) for i, v in enumerate(vals)}
        mkt_current[name] = round(vals[-1], 1)
        mkt_changes[name] = {
            "1d": round(vals[-1] - vals[-2], 1) if len(vals) > 1 else None,
            "1w": round(vals[-1] - vals[-6], 1) if len(vals) > 5 else None,
            "1m": round(vals[-1] - vals[-22], 1) if len(vals) > 21 else None,
        }

    # 거래 샘플 (TRACE 형식)
    sources = ["D2D","CT","AT","BN","ML","GS","JP","MS","CS","UB"]
    trades = []
    for b in bonds[:15]:
        for _ in range(3):
            dd = datetime.today() - timedelta(days=np.random.randint(0, 7))
            trd_sp = round(b["spread"] + (np.random.random() - 0.5) * 10, 1) if b["spread"] else None
            yas_sp = b["spread"]
            sp_chg = round(trd_sp - yas_sp, 1) if trd_sp and yas_sp else None
            trades.append({
                "date": dd.strftime("%Y-%m-%d"),
                "bond": b["label"],
                "price": round(b["price"] + (np.random.random() - 0.5) * 2, 3),
                "ytm": round(b["ytm"] + (np.random.random() - 0.5) * 0.3, 3),
                "trd_spread": trd_sp,
                "yas_spread": yas_sp,
                "spread_chg": sp_chg,
                "volume": int(np.random.randint(1, 50) * 100000),
                "source": np.random.choice(sources),
            })
    trades.sort(key=lambda x: x["date"], reverse=True)

    # ETF 샘플
    etf_names = list(ETF_TICKERS.values())
    etf_bases = {"VCSH": 45, "VCIT": 75, "LQD": 95, "VCLT": 130, "IEAC": 85, "HYG": 320, "EMB": 280, "EMLC": 350}
    etf_vols = {"VCSH": 2, "VCIT": 3, "LQD": 3, "VCLT": 4, "IEAC": 3, "HYG": 10, "EMB": 8, "EMLC": 12}
    etf_ts = {}
    etf_current = {}
    etf_changes = {}
    for name in etf_names:
        base = etf_bases.get(name, 100)
        vol = etf_vols.get(name, 3)
        vals = [base]
        for i in range(1, len(mkt_dates)):
            vals.append(max(base * 0.6, min(base * 1.5, vals[-1] + (np.random.random() - 0.48) * vol)))
        etf_ts[name] = {mkt_date_strs[i]: round(float(v), 1) for i, v in enumerate(vals)}
        etf_current[name] = round(vals[-1], 1)
        etf_changes[name] = {
            "1d": round(vals[-1] - vals[-2], 1) if len(vals) > 1 else None,
            "1w": round(vals[-1] - vals[-6], 1) if len(vals) > 5 else None,
            "1m": round(vals[-1] - vals[-22], 1) if len(vals) > 21 else None,
            "3m": round(vals[-1] - vals[-63], 1) if len(vals) > 63 else None,
            "6m": round(vals[-1] - vals[0], 1) if len(vals) > 1 else None,
        }

    return {
        "generated": datetime.today().strftime("%Y-%m-%d %H:%M"),
        "bondCount": len(bonds),
        "bondIsins": BOND_ISINS,
        "bonds": bonds,
        "allBonds": all_bonds,
        "spreadHistory": {"dates": date_strs, "series": series},
        "ustCurve": {
            "tenors": list(UST_TICKERS.keys()),
            "years": [1/12,.25,.5,1,2,3,5,7,10,20,30],
            "now": [5.31,5.28,5.12,4.85,4.62,4.52,4.41,4.38,4.32,4.51,4.55],
            "ago3m": [5.05,5.10,4.95,4.68,4.44,4.33,4.25,4.22,4.18,4.37,4.43],
        },
        "marketTimeseries": mkt_ts,
        "marketCurrent": mkt_current,
        "marketChanges": mkt_changes,
        "etfTimeseries": etf_ts,
        "etfCurrent": etf_current,
        "etfChanges": etf_changes,
        "trades": trades,
        "newIssues": merge_new_issues(load_prel_data() or [
            {"date":"05/19","cor":"US","ighy":"IG","ticker":"F 6.467 05/22/36","issuer":"Ford Motor Credit Co LLC","sp":None,"ipt":"T + 205 Area","fg":"T + 180","nic":25.0,"amt":"1.00B","ccy":"USD","coupon":6.467,"tenor":"10.0","status":"Priced","rank":"Sr Unsecured","reg":"144a"},
            {"date":"05/19","cor":"FR","ighy":"IG","ticker":"SOCGEN 5.371 05/27/32","issuer":"Societe Generale SA","sp":None,"ipt":"T + 125 Area","fg":"T + 105","nic":20.0,"amt":"1.25B","ccy":"USD","coupon":5.371,"tenor":"6-NC5","status":"Priced","rank":"Sr Non Preferred","reg":"144a"},
            {"date":"05/19","cor":"CA","ighy":"IG","ticker":"CCDJ 5.021 05/27/31","issuer":"Fed Caisses Desjardins","sp":None,"ipt":"T + 95-100","fg":"T + 70","nic":27.5,"amt":"750.00M","ccy":"USD","coupon":5.021,"tenor":"5.0","status":"Priced","rank":"Sr Unsecured","reg":"144a"},
            {"date":"05/19","cor":"US","ighy":"IG","ticker":"ECL 5.35 06/15/36","issuer":"Ecolab Inc","sp":None,"ipt":"T + 95 Area","fg":"T + 73","nic":22.0,"amt":"1.40B","ccy":"USD","coupon":5.35,"tenor":"10.0","status":"Priced","rank":"Sr Unsecured","reg":"SEC"},
            {"date":"05/19","cor":"US","ighy":"IG","ticker":"ECL 4.8 06/15/31","issuer":"Ecolab Inc","sp":None,"ipt":"T + 75 Area","fg":"T + 53","nic":22.0,"amt":"900.00M","ccy":"USD","coupon":4.8,"tenor":"5.0","status":"Priced","rank":"Sr Unsecured","reg":"SEC"},
            {"date":"05/19","cor":"US","ighy":"IG","ticker":"PRU 5 05/27/31","issuer":"Pricoa Global Funding I","sp":None,"ipt":"T + 90 Area","fg":"T + 68","nic":22.0,"amt":"450.00M","ccy":"USD","coupon":5.0,"tenor":"5.0","status":"Priced","rank":"Secured","reg":"144a"},
            {"date":"05/19","cor":"US","ighy":"IG","ticker":"MBGL 6.05 06/15/36","issuer":"Mobility Global Inc","sp":None,"ipt":"T + 170 Area","fg":"T + 170 Area","nic":None,"amt":"700.00M","ccy":"USD","coupon":6.05,"tenor":"10.0","status":"Priced","rank":"Sr Unsecured","reg":"SEC"},
            {"date":"05/19","cor":"US","ighy":"HY","ticker":"RRD 0 5.00Y","issuer":"RR Donnelley & Sons Co","sp":None,"ipt":"11%","fg":None,"nic":None,"amt":"750.00M","ccy":"USD","coupon":None,"tenor":"5-NC2","status":"Talk","rank":"Sr Unsecured","reg":"144a"},
        ]),
        "bdcCurrent":    {meta["short"]: {"price": None, "pb": None, "nav": None} for meta in BDC_TICKERS.values()},
        "bdcChanges":    {meta["short"]: {"1d": None, "1w": None, "1m": None} for meta in BDC_TICKERS.values()},
        "bdcTimeseries": {},
        "bdcMeta":       {meta["short"]: {"name": meta["name"], "color": meta["color"]} for meta in BDC_TICKERS.values()},
    }


# ══════════════════════════════════════════════
# 4. HTML 템플릿 생성
# ══════════════════════════════════════════════
def generate_html(data):
    """데이터를 받아 standalone HTML 대시보드 문자열을 반환"""

    data_json = json.dumps(data, ensure_ascii=False)

    html = f'''<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>IG Credit Spread Dashboard</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:opsz,wght@9..40,300;9..40,500;9..40,700&family=JetBrains+Mono:wght@400;500;600&display=swap');
:root {{
  --bg:#0b0f1a;--sf:#111827;--sf2:#1a2235;--sf3:#222d42;--bd:#2a3550;
  --tx:#e8ecf4;--tx2:#8b98b0;--tx3:#5c6b85;
  --ac:#3b82f6;--acd:#1e40af;--gn:#10b981;--rd:#ef4444;--am:#f59e0b;--cy:#06b6d4;--pu:#8b5cf6;
  --ft:'DM Sans',sans-serif;--mn:'JetBrains Mono',monospace;
}}
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:var(--ft);background:var(--bg);color:var(--tx);min-height:100vh;-webkit-font-smoothing:antialiased}}
.hd{{background:linear-gradient(135deg,#0f1729,#162040);border-bottom:1px solid var(--bd);padding:16px 32px;display:flex;justify-content:space-between;align-items:center;position:sticky;top:0;z-index:100;backdrop-filter:blur(12px)}}
.hd h1{{font-size:18px;font-weight:700;letter-spacing:-.3px}}
.hd .sub{{font-size:11px;color:var(--tx3);margin-top:2px;font-family:var(--mn)}}
.hd .st{{display:flex;align-items:center;gap:6px;font-size:12px;color:var(--gn)}}
.hd .st::before{{content:'';width:7px;height:7px;border-radius:50%;background:var(--gn);animation:pulse 2s infinite}}
@keyframes pulse{{0%,100%{{opacity:1}}50%{{opacity:.4}}}}
.tabs{{display:flex;gap:2px;padding:12px 32px 0;background:var(--bg);border-bottom:1px solid var(--bd)}}
.tab{{padding:10px 20px;font-size:13px;font-weight:500;color:var(--tx3);cursor:pointer;border:1px solid transparent;border-bottom:none;border-radius:8px 8px 0 0;transition:all .2s;user-select:none}}
.tab:hover{{color:var(--tx2);background:var(--sf)}}
.tab.active{{color:var(--ac);background:var(--sf);border-color:var(--bd);position:relative}}
.tab.active::after{{content:'';position:absolute;bottom:-1px;left:0;right:0;height:2px;background:var(--ac)}}
.ct{{padding:20px 32px 40px;max-width:1440px;margin:0 auto}}
.tp{{display:none}}.tp.active{{display:block;animation:fi .3s}}
@keyframes fi{{from{{opacity:0;transform:translateY(6px)}}to{{opacity:1;transform:none}}}}
.ms{{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px;margin-bottom:20px}}
.mc{{background:var(--sf);border:1px solid var(--bd);border-radius:10px;padding:16px;transition:border-color .2s}}
.mc:hover{{border-color:var(--acd)}}
.mc .lb{{font-size:11px;color:var(--tx3);text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px}}
.mc .vl{{font-size:24px;font-weight:700;font-family:var(--mn)}}
.mc .dl{{font-size:11px;margin-top:6px;font-family:var(--mn)}}
.mc .dl.up{{color:var(--rd)}}.mc .dl.dn{{color:var(--gn)}}
.cd{{background:var(--sf);border:1px solid var(--bd);border-radius:10px;padding:20px;margin-bottom:16px}}
.ct2{{font-size:14px;font-weight:600;margin-bottom:14px;display:flex;align-items:center;gap:8px}}
.ib{{background:rgba(59,130,246,.08);border:1px solid rgba(59,130,246,.2);border-radius:8px;padding:10px 16px;font-size:12px;color:var(--ac);margin-bottom:16px}}
.g2{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:16px}}
.g3{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;margin-bottom:16px}}
@media(max-width:900px){{.g2,.g3,.ig{{grid-template-columns:1fr}}}}
.tw{{overflow-x:auto;border-radius:8px;border:1px solid var(--bd)}}
table{{width:100%;border-collapse:collapse;font-size:12px}}
thead th{{padding:10px 12px;text-align:left;font-size:11px;font-weight:600;color:var(--tx3);text-transform:uppercase;letter-spacing:.4px;background:var(--sf2);border-bottom:2px solid var(--bd);position:sticky;top:0;cursor:pointer;user-select:none;white-space:nowrap}}
thead th:hover{{color:var(--tx)}}thead th.so{{color:var(--ac)}}
tbody td{{padding:9px 12px;border-bottom:1px solid rgba(42,53,80,.5);font-family:var(--mn);font-size:11.5px;white-space:nowrap}}
tbody tr:hover{{background:rgba(59,130,246,.04)}}
td.tc{{font-family:var(--ft)}}td.nm{{text-align:right}}td.ps{{color:var(--rd)}}td.ng{{color:var(--gn)}}td.sp{{font-weight:600}}
.bg{{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:12px}}
.bt{{padding:6px 14px;font-size:12px;font-family:var(--ft);font-weight:500;border:1px solid var(--bd);border-radius:6px;background:var(--sf2);color:var(--tx2);cursor:pointer;transition:all .15s}}
.bt:hover{{border-color:var(--acd);color:var(--tx);background:var(--sf3)}}
.bt.pr{{background:var(--ac);color:#fff;border-color:var(--ac)}}.bt.pr:hover{{background:#2563eb}}
.fl{{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:14px}}
.fl select,.fl input{{padding:7px 12px;font-size:12px;font-family:var(--ft);background:var(--sf2);color:var(--tx);border:1px solid var(--bd);border-radius:6px;outline:none}}
.fl select:focus,.fl input:focus{{border-color:var(--ac)}}
.fl label{{font-size:12px;font-weight:500;color:var(--tx3)}}
.ig{{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:20px}}
.ic{{background:var(--sf);border:1px solid var(--bd);border-radius:10px;padding:14px}}
.ic .it{{font-size:13px;font-weight:600;margin-bottom:4px}}
.ic .is{{font-size:11px;color:var(--tx3);margin-bottom:10px}}
.mi{{border-left:3px solid var(--pu);padding:12px 16px;margin-bottom:10px;background:var(--sf2);border-radius:0 6px 6px 0}}
.mi .mh{{display:flex;justify-content:space-between;margin-bottom:6px;align-items:center}}
.mi .mt{{font-size:10px;background:var(--pu);color:white;padding:2px 8px;border-radius:4px;font-weight:600}}
.mi .mtl{{font-size:13px;font-weight:600;margin-left:8px}}
.mi .md{{font-size:11px;color:var(--tx3);font-family:var(--mn)}}
.mi .mb{{font-size:12px;color:var(--tx2);white-space:pre-wrap;line-height:1.6}}
textarea,.ti{{width:100%;padding:10px 14px;font-size:13px;font-family:var(--ft);background:var(--sf2);color:var(--tx);border:1px solid var(--bd);border-radius:6px;outline:none;resize:vertical}}
textarea:focus,.ti:focus{{border-color:var(--ac)}}
::-webkit-scrollbar{{width:6px;height:6px}}::-webkit-scrollbar-track{{background:transparent}}::-webkit-scrollbar-thumb{{background:var(--bd);border-radius:3px}}
.ch{{position:relative;width:100%}}.ch canvas{{display:block}}
.en{{position:fixed;bottom:20px;right:20px;background:var(--gn);color:#fff;padding:10px 20px;border-radius:8px;font-size:13px;font-weight:500;opacity:0;transform:translateY(10px);transition:all .3s;pointer-events:none}}.en.show{{opacity:1;transform:none}}
/* Legend dot for issuer curves */
.lg{{display:flex;gap:16px;flex-wrap:wrap;margin-bottom:8px;font-size:11px;color:var(--tx2)}}
.lg span::before{{content:'';display:inline-block;width:8px;height:8px;border-radius:50%;margin-right:4px;vertical-align:middle}}
.lg .owned::before{{background:currentColor}}.lg .extra::before{{background:transparent;border:2px solid currentColor}}
</style>
</head>
<body>
<div class="hd"><div><h1>IG Credit Spread Dashboard</h1><div class="sub" id="hdr-sub"></div></div><div class="st" id="hdr-status"></div></div>
<div class="tabs">
  <div class="tab active" data-tab="t1">🎯 보유 채권</div>
  <div class="tab" data-tab="t2">📊 크레딧 지표</div>
  <div class="tab" data-tab="t3">🏢 발행자별 곡선</div>
  <div class="tab" data-tab="t4">💼 거래 내역</div>
  <div class="tab" data-tab="t6">🆕 최근 발행</div>
  <div class="tab" data-tab="t5">📝 리서치 메모</div>
  <div class="tab" data-tab="t7">⚙️ 종목 관리</div>
</div>
<div class="ct">
<!-- TAB 1 -->
<div class="tp active" id="t1">
  <div class="ms" id="m1"></div>
  <div class="ib" id="ss">📌 종목을 선택하면 평균 통계가 표시됩니다</div>
  <div class="cd">
    <div class="bg"><span style="font-size:12px;color:var(--tx3);font-weight:600;line-height:28px">빠른 선택:</span>
      <button class="bt" onclick="SA()">전체</button><button class="bt" onclick="SN()">해제</button>
      <button class="bt" onclick="ST()">상위 10</button><button class="bt" onclick="SB()">하위 10</button>
      <span style="width:1px;height:20px;background:var(--bd);margin:0 4px"></span>
      <span style="font-size:11px;color:var(--tx3);line-height:28px">지역:</span>
      <button class="bt" onclick="SR('미국')">미국</button><button class="bt" onclick="SR('유럽')">유럽</button><button class="bt" onclick="SR('아시아/기타')">아시아</button>
      <span style="width:1px;height:20px;background:var(--bd);margin:0 4px"></span>
      <span style="font-size:11px;color:var(--tx3);line-height:28px">섹터:</span>
      <button class="bt" onclick="SS('금융기관')">금융</button><button class="bt" onclick="SS('기술')">기술</button><button class="bt" onclick="SS('에너지')">에너지</button><button class="bt" onclick="SS('산업')">산업</button><button class="bt" onclick="SS('유틸리티')">유틸</button><button class="bt" onclick="SS('자유 소비재')">소비</button><button class="bt" onclick="SS('원자재')">원자재</button><button class="bt" onclick="SS('통신')">통신</button>
      <span style="width:1px;height:20px;background:var(--bd);margin:0 4px"></span>
      <span style="font-size:11px;color:var(--tx3);line-height:28px">만기:</span>
      <button class="bt" onclick="SM2('1-3년')">1-3Y</button><button class="bt" onclick="SM2('3-5년')">3-5Y</button><button class="bt" onclick="SM2('5-7년')">5-7Y</button><button class="bt" onclick="SM2('7-10년')">7-10Y</button><button class="bt" onclick="SM2('10-20년')">10Y+</button>
    </div>
    <div class="fl">
      <label>섹터</label><select id="fs" onchange="AF()"><option value="ALL">전체</option></select>
      <label>만기</label><select id="fm" onchange="AF()"><option value="ALL">전체</option></select>
      <label>지역</label><select id="fr" onchange="AF()"><option value="ALL">전체</option></select>
      <label>검색</label><input type="text" id="fx" placeholder="채권명 검색..." oninput="AF()" style="width:200px">
    </div>
    <div id="fc" style="font-size:11px;color:var(--tx3);margin-bottom:6px"></div>
  </div>
  <div class="cd"><div class="ct2">📈 G-spread 추이 (bps, 3개월)</div><div class="ch" style="height:320px"><canvas id="c1"></canvas></div></div>
  <div class="g2">
    <div class="cd"><div class="ct2">📐 듀레이션 vs 크레딧 스프레드</div><div class="ch" style="height:280px"><canvas id="c2"></canvas></div></div>
    <div class="cd"><div class="ct2">📊 현재 G-spread 비교</div><div class="ch" style="height:280px"><canvas id="c3"></canvas></div></div>
  </div>

  <!-- 포트폴리오 테이블 -->
  <div class="cd">
    <div class="ct2">📋 종목별 현황 + 포트폴리오 손익 <span style="font-size:11px;color:var(--tx3);font-weight:400;margin-left:8px">컬럼 클릭=정렬 | 행 클릭=선택 | 수량은 자동 저장됩니다</span></div>
    <div class="fl">
      <label>섹터</label><select id="fs2" onchange="RT()"><option value="ALL">전체</option></select>
      <label>만기</label><select id="fm2" onchange="RT()"><option value="ALL">전체</option></select>
      <label>지역</label><select id="fr2" onchange="RT()"><option value="ALL">전체</option></select>
      <label>검색</label><input type="text" id="fx2" placeholder="채권명 검색..." oninput="RT()" style="width:200px">
      <span style="width:1px;height:20px;background:var(--bd);margin:0 6px"></span>
      <label>손익 기준</label>
      <select id="pnl-scenario" onchange="RT()">
        <option value="1d">1D</option>
        <option value="1w">1W</option>
        <option value="1m" selected>1M</option>
        <option value="3m">3M</option>
      </select>
      <button class="bt pr" onclick="EX()">📥 CSV</button>
    </div>
    <div id="pnl-summary" class="ib" style="margin-bottom:6px"></div>
    <div class="tw" style="max-height:600px;overflow-y:auto"><table><thead><tr id="th"></tr></thead><tbody id="tb"></tbody></table></div>
  </div>
</div>
<!-- TAB 2 -->
<div class="tp" id="t2">
  <div class="ib">📊 크레딧 시장 전반의 위험도와 투자심리를 보여주는 핵심 지표</div>
  <div class="ms" id="m2"></div>
  <div class="cd"><div class="ct2">📉 크레딧 스프레드 인덱스 (6개월)</div><div class="ch" style="height:300px"><canvas id="c4"></canvas></div></div>
  <div class="g2">
    <div class="cd"><div class="ct2">⚡ VIX</div><div class="ch" style="height:240px"><canvas id="c5"></canvas></div></div>
    <div class="cd"><div class="ct2">🔔 MOVE</div><div class="ch" style="height:240px"><canvas id="c6"></canvas></div></div>
  </div>
  <div class="cd" id="bdcSection">
    <div class="ct2">📊 주요 BDC 현황 (상장 BDC)</div>
    <div id="bdcCards" style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:14px;"></div>
    <div class="g2">
      <div style="background:var(--bg1);border:1px solid var(--bd);border-radius:8px;padding:14px;">
        <div style="font-size:12px;font-weight:600;color:var(--tx2);margin-bottom:8px;">NAV 대비 할인율 (%)</div>
        <div style="height:190px;position:relative;"><canvas id="bdcBarChart"></canvas></div>
      </div>
      <div style="background:var(--bg1);border:1px solid var(--bd);border-radius:8px;padding:14px;">
        <div style="font-size:12px;font-weight:600;color:var(--tx2);margin-bottom:8px;">주가 추이 (시작=100, 6M)</div>
        <div style="height:190px;position:relative;"><canvas id="bdcLineChart"></canvas></div>
      </div>
    </div>
  </div>
  <div class="cd"><div class="ct2">💳 ETF 크레딧 스프레드 현황 (OAS, bps)</div>
    <div class="tw"><table><thead><tr><th>ETF</th><th>현재</th><th>1D</th><th>1W</th><th>1M</th><th>3M</th><th>6M</th></tr></thead><tbody id="etfTbl"></tbody></table></div></div>
  <div class="g2">
    <div class="cd"><div class="ct2">📈 IG ETF (VCSH / VCIT / LQD / VCLT / IEAC)</div><div class="ch" style="height:260px"><canvas id="c9"></canvas></div></div>
    <div class="cd"><div class="ct2">📉 HY / EM ETF (HYG / EMB / EMLC)</div><div class="ch" style="height:260px"><canvas id="c10"></canvas></div></div>
  </div>
</div>
<!-- TAB 3 -->
<div class="tp" id="t3">
  <div class="ib">🏢 발행자별 개별 수익률 곡선 — ● 보유  ○ 비보유</div>
  <div class="ig" id="ic"></div>
  <div class="cd"><div class="ct2">🏭 섹터 평균 곡선</div><div class="ch" style="height:300px"><canvas id="c7"></canvas></div></div>
</div>
<!-- TAB 4 -->
<div class="tp" id="t4">
  <div class="ib">💼 TRACE 거래 데이터 (최근 7일)</div>
  <div class="cd"><div class="ct2">💹 거래 내역</div>
    <div class="tw" style="max-height:600px;overflow-y:auto">
      <table><thead><tr><th>날짜</th><th>채권</th><th>소스</th><th>가격</th><th>YTM(%)</th><th>거래 Spread</th><th>YAS Spread</th><th>변동</th><th>거래량</th><th>사이즈</th></tr></thead><tbody id="t4b"></tbody></table></div></div>
</div>
<!-- TAB 6: New Issues -->
<div class="tp" id="t6">
  <div class="ib">🆕 Bloomberg PREL 워크시트 기반 최근 발행 내역 — <span id="niCount"></span></div>
  <div class="cd">
    <div class="fl">
      <label>IG/HY</label><select id="nf1" onchange="RNI()"><option value="ALL">전체</option><option value="IG">IG</option><option value="HY">HY</option></select>
      <label>Status</label><select id="nf2" onchange="RNI()"><option value="ALL">전체</option><option value="Priced">Priced</option><option value="Talk">Talk</option><option value="Launch">Launch</option></select>
      <label>CoR</label><select id="nf3" onchange="RNI()"><option value="ALL">전체</option></select>
      <label>검색</label><input type="text" id="nf4" placeholder="발행자/티커 검색..." oninput="RNI()" style="width:200px">
    </div>
    <div class="tw" style="max-height:600px;overflow-y:auto">
      <table><thead><tr>
        <th>날짜</th><th>CoR</th><th>IG/HY</th><th>티커</th><th>발행자</th>
        <th>IPT</th><th>Final</th><th>NIC</th><th>금액</th><th>통화</th>
        <th>쿠폰</th><th>테너</th><th>Status</th><th>구조</th><th>Reg</th>
      </tr></thead><tbody id="nib"></tbody></table>
    </div>
  </div>
</div>
<!-- TAB 5 -->
<div class="tp" id="t5">
  <div class="ib">📝 리서치 핵심 내용을 메모해두세요. 브라우저에 자동 저장됩니다.</div>
  <div class="cd">
    <div style="display:flex;gap:12px;flex-wrap:wrap;margin-bottom:12px">
      <div style="flex:1;min-width:180px"><label style="font-size:12px;color:var(--tx3);display:block;margin-bottom:4px">발행자/섹터</label><input class="ti" id="rk" placeholder="예: HSBC"></div>
      <div style="flex:2;min-width:200px"><label style="font-size:12px;color:var(--tx3);display:block;margin-bottom:4px">제목</label><input class="ti" id="rt" placeholder="리서치 제목"></div></div>
    <label style="font-size:12px;color:var(--tx3);display:block;margin-bottom:4px">요약 / 메모</label>
    <textarea id="rc" rows="5" placeholder="핵심 내용을 메모..."></textarea>
    <div style="margin-top:12px;display:flex;gap:12px;align-items:center">
      <button class="bt pr" onclick="SM()">저장</button><span id="rs" style="font-size:12px;color:var(--tx3)"></span></div>
  </div>
  <div class="cd"><div class="ct2">🗂 저장된 리서치 메모</div><div id="rl"></div></div>
</div>
</div>
<!-- TAB 7: Bond Management -->
<div class="tp" id="t7">
  <div class="ib">⚙️ 보유 채권 ISIN 목록을 관리합니다. 변경 후 다운로드 → 폴더에 저장 → 다음 실행 시 반영됩니다.</div>
  <div class="cd">
    <div class="ct2">➕ 종목 추가</div>
    <div class="fl">
      <input type="text" id="addIsin" placeholder="ISIN 또는 Bloomberg ID 입력 (예: US037833FA32)" style="width:350px" class="ti">
      <button class="bt pr" onclick="addBond()">추가</button>
    </div>
    <div id="addMsg" style="font-size:12px;color:var(--tx3);margin-top:6px"></div>
  </div>
  <div class="cd">
    <div class="ct2">📋 현재 종목 목록 (<span id="bondListCount"></span>개)
      <span style="font-size:11px;color:var(--tx3);font-weight:400;margin-left:12px">삭제할 종목 체크 후 아래 삭제 버튼 클릭</span>
    </div>
    <div class="fl">
      <button class="bt" onclick="selectAllBonds()">전체 선택</button>
      <button class="bt" onclick="deselectAllBonds()">선택 해제</button>
      <button class="bt" style="border-color:var(--rd);color:var(--rd)" onclick="deleteSelected()">선택 삭제</button>
      <span style="width:1px;height:20px;background:var(--bd);margin:0 8px"></span>
      <button class="bt pr" onclick="downloadBondList()">📥 bond_list.json 다운로드</button>
      <span id="dlMsg" style="font-size:12px;color:var(--gn);margin-left:8px"></span>
    </div>
    <div class="tw" style="max-height:500px;overflow-y:auto">
      <table><thead><tr><th style="width:40px"></th><th>ISIN / ID</th><th>종목명</th><th>발행자</th><th>섹터</th></tr></thead><tbody id="blBody"></tbody></table>
    </div>
  </div>
  <div class="cd" style="background:rgba(239,68,68,.05);border-color:rgba(239,68,68,.2)">
    <div class="ct2" style="color:var(--rd)">⚠️ 적용 방법</div>
    <div style="font-size:13px;color:var(--tx2);line-height:1.8">
      1. 위에서 종목을 추가/삭제한 후 <b>bond_list.json 다운로드</b> 클릭<br>
      2. 다운로드된 파일을 <b>credit-dashboard</b> 폴더에 저장 (덮어쓰기)<br>
      3. <b>update_dashboard.bat</b> 실행하면 다음 업데이트부터 반영
    </div>
  </div>
</div>
</div>
<div class="en" id="en">✅ CSV 파일이 다운로드됩니다</div>
<script>
// ═══ DATA ═══
const D = {data_json};

const CL = ['#3b82f6','#10b981','#ef4444','#8b5cf6','#f59e0b','#06b6d4','#ec4899','#14b8a6','#f97316','#6366f1','#84cc16','#e11d48','#0ea5e9','#a855f7','#22c55e'];

// Header
document.getElementById('hdr-sub').textContent = `G-spread | 3M | ${{D.generated}} | ${{D.bondCount}} bonds`;
document.getElementById('hdr-status').textContent = D.bonds[0]?.issuer?.includes?.('샘플') ? 'Sample Data' : 'Bloomberg Data (' + D.generated.split(' ')[0] + ')';

// Chart.js defaults
Chart.defaults.color='#8b98b0';Chart.defaults.borderColor='rgba(42,53,80,.4)';
Chart.defaults.font.family="'DM Sans',sans-serif";Chart.defaults.font.size=11;
Chart.defaults.plugins.legend.labels.usePointStyle=true;Chart.defaults.plugins.legend.labels.pointStyleWidth=10;
Chart.defaults.plugins.legend.labels.padding=12;Chart.defaults.animation.duration=600;

// ═══ Tab switching ═══
document.querySelectorAll('.tab').forEach(t=>t.addEventListener('click',()=>{{
  document.querySelectorAll('.tab').forEach(x=>x.classList.remove('active'));
  document.querySelectorAll('.tp').forEach(x=>x.classList.remove('active'));
  t.classList.add('active');document.getElementById(t.dataset.tab).classList.add('active');
  if(t.dataset.tab==='t2')R2();if(t.dataset.tab==='t3')R3();
  if(t.dataset.tab==='t4')R4();if(t.dataset.tab==='t6')R6();if(t.dataset.tab==='t5')RL();if(t.dataset.tab==='t7')R7();
}}));

// ═══ TAB 1 ═══
const bonds=D.bonds;
let sel=new Set(bonds.slice(0,12).map(b=>b.label));
let sc='spread',sd=-1;
let ch1,ch2,ch3;

function GS(){{return GF().filter(b=>sel.has(b.label))}}
function GF(){{
  var s=document.getElementById('fs').value,m=document.getElementById('fm').value,
      r=document.getElementById('fr').value,x=document.getElementById('fx').value.toLowerCase();
  return bonds.filter(function(b){{
    if(s!=='ALL'&&b.sector!==s)return false;
    if(m!=='ALL'&&b.matBucket!==m)return false;
    if(r!=='ALL'&&b.region!==r)return false;
    if(x&&!b.label.toLowerCase().includes(x)&&!(b.issuer||'').toLowerCase().includes(x))return false;
    return true;
  }});
}}
function GF2(){{
  var s=document.getElementById('fs2').value,m=document.getElementById('fm2').value,
      r=document.getElementById('fr2').value,x=document.getElementById('fx2').value.toLowerCase();
  return bonds.filter(function(b){{
    if(s!=='ALL'&&b.sector!==s)return false;
    if(m!=='ALL'&&b.matBucket!==m)return false;
    if(r!=='ALL'&&b.region!==r)return false;
    if(x&&!b.label.toLowerCase().includes(x)&&!(b.issuer||'').toLowerCase().includes(x))return false;
    return true;
  }});
}}
function resetFilters(){{
  document.getElementById('fs').value='ALL';
  document.getElementById('fm').value='ALL';
  document.getElementById('fr').value='ALL';
  document.getElementById('fx').value='';
}}
function SA(){{resetFilters();bonds.forEach(b=>sel.add(b.label));R1()}}
function SN(){{resetFilters();sel.clear();R1()}}
function ST(){{resetFilters();sel.clear();[...bonds].sort((a,b)=>(b.spread||0)-(a.spread||0)).slice(0,10).forEach(b=>sel.add(b.label));R1()}}
function SB(){{resetFilters();sel.clear();[...bonds].sort((a,b)=>(a.spread||0)-(b.spread||0)).slice(0,10).forEach(b=>sel.add(b.label));R1()}}
function SR(r){{resetFilters();sel.clear();bonds.filter(b=>b.region===r).forEach(b=>sel.add(b.label));R1()}}
function SS(s){{resetFilters();sel.clear();var matched=bonds.filter(function(b){{return b.sector===s}});console.log('SS('+s+'): matched='+matched.length);matched.forEach(function(b){{sel.add(b.label)}});console.log('sel.size='+sel.size);R1()}}
function SM2(m){{resetFilters();sel.clear();bonds.filter(b=>b.matBucket===m).forEach(b=>sel.add(b.label));R1()}}
function AF(){{R1()}}
function UC(){{
  const f=GF(),s2=f.filter(b=>sel.has(b.label));
  document.getElementById('fc').textContent=`선택: ${{sel.size}}개 / 필터: ${{f.length}}개 / 전체: ${{bonds.length}}개`;
}}

function RM(){{
  const s2=GS(),avg=s2.length?s2.reduce((a,b)=>a+(b.spread||0),0)/s2.length:0;
  const mx=s2.length?Math.max(...s2.map(b=>b.spread||0)):0;
  const mn=s2.length?Math.min(...s2.map(b=>b.spread||0)):0;
  const u10=D.ustCurve.now[8];
  const mk=(l,v)=>`<div class="mc"><div class="lb">${{l}}</div><div class="vl">${{v}}</div></div>`;
  document.getElementById('m1').innerHTML=[mk('모니터링 종목',bonds.length+'개'),mk('평균 G-spread',avg.toFixed(0)+' bps'),mk('최고 / 최저',mx+' / '+mn+' bps'),mk('UST 10Y',(u10||0).toFixed(2)+'%')].join('');
}}

function RS(){{
  const s2=GS();
  if(!s2.length){{document.getElementById('ss').innerHTML='📌 종목을 선택하면 평균 통계가 표시됩니다';return}}
  const avg=(a,f)=>a.reduce((s,b)=>s+(f(b)||0),0)/a.length;
  const fc=(v)=>v!=null?(v>=0?'+':'')+v.toFixed(1):'-';
  const cc=(v)=>v>0?'color:var(--rd)':v<0?'color:var(--gn)':'';
  const as=avg(s2,b=>b.spread).toFixed(0),ay=avg(s2,b=>b.ytm).toFixed(2);
  const a1=avg(s2,b=>b.chg1d),aw=avg(s2,b=>b.chg1w),am=avg(s2,b=>b.chg1m);
  document.getElementById('ss').innerHTML=`<span style="font-weight:600;color:var(--ac)">📌 선택 ${{s2.length}}개 평균</span>
    <span style="margin-left:16px"><span style="color:var(--tx3)">G-spread:</span> <b>${{as}} bps</b></span>
    <span style="margin-left:16px"><span style="color:var(--tx3)">YTM:</span> <b>${{ay}}%</b></span>
    <span style="margin-left:16px"><span style="color:var(--tx3)">1D:</span> <b style="${{cc(a1)}}">${{fc(a1)}}</b></span>
    <span style="margin-left:16px"><span style="color:var(--tx3)">1W:</span> <b style="${{cc(aw)}}">${{fc(aw)}}</b></span>
    <span style="margin-left:16px"><span style="color:var(--tx3)">1M:</span> <b style="${{cc(am)}}">${{fc(am)}}</b></span>`;
}}

function RC1(){{
  const s2=GS().slice(0,15);const ctx=document.getElementById('c1').getContext('2d');
  if(ch1)ch1.destroy();
  const H=D.spreadHistory;
  ch1=new Chart(ctx,{{type:'line',data:{{labels:H.dates,datasets:s2.map((b,i)=>({{
    label:b.label.length>25?b.label.slice(0,22)+'...':b.label,
    data:H.series[b.label]||[],borderColor:CL[i%CL.length],borderWidth:1.5,pointRadius:0,pointHitRadius:8,tension:.3
  }}))
  }},options:{{responsive:true,maintainAspectRatio:false,interaction:{{mode:'index',intersect:false}},
    plugins:{{legend:{{display:true,position:'top'}},tooltip:{{callbacks:{{label:c=>c.dataset.label+': '+c.parsed.y?.toFixed(0)+' bps'}}}}}},
    scales:{{x:{{ticks:{{maxTicksLimit:10,font:{{size:10}}}},grid:{{display:false}}}},y:{{title:{{display:true,text:'bps'}},grid:{{color:'rgba(42,53,80,.3)'}}}}}}
  }}}});
}}

function RC2(){{
  const s2=GS();const ctx=document.getElementById('c2').getContext('2d');if(ch2)ch2.destroy();
  const selBonds=s2.filter(b=>b.duration&&b.spread);
  if(!selBonds.length){{
    ch2=new Chart(ctx,{{type:'scatter',data:{{datasets:[]}},options:{{responsive:true,maintainAspectRatio:false}}}});
    return;
  }}
  const rc=r=>{{if(!r)return'#6b7280';if(r.startsWith('AA'))return'#10b981';if(r.startsWith('A'))return'#f59e0b';return'#ef4444'}};
  ch2=new Chart(ctx,{{type:'scatter',
    data:{{datasets:[{{
      label:'선택 채권 G-spread',
      data:selBonds.map(b=>({{x:b.duration,y:b.spread}})),
      backgroundColor:selBonds.map(b=>rc(b.rating)),
      pointRadius:7,pointHoverRadius:10,showLine:false,
    }}]}},
    options:{{responsive:true,maintainAspectRatio:false,
      plugins:{{
        legend:{{display:false}},
        tooltip:{{callbacks:{{
          label:function(c){{
            const b=selBonds[c.dataIndex];
            if(!b)return'';
            return b.label+' | '+b.duration.toFixed(1)+'Y | '+b.spread.toFixed(0)+'bps | '+b.rating;
          }}
        }}}}
      }},
      scales:{{
        x:{{title:{{display:true,text:'듀레이션(년)'}},min:0,grid:{{color:'rgba(42,53,80,.3)'}}}},
        y:{{title:{{display:true,text:'G-spread (bps)'}},grid:{{color:'rgba(42,53,80,.3)'}}}}
      }}
    }}
  }});
}}

function RC3(){{
  const s2=GS().sort((a,b)=>(a.spread||0)-(b.spread||0)).slice(0,20);
  const ctx=document.getElementById('c3').getContext('2d');if(ch3)ch3.destroy();
  ch3=new Chart(ctx,{{type:'bar',data:{{labels:s2.map(b=>b.label.length>20?b.label.slice(0,18)+'...':b.label),
    datasets:[{{data:s2.map(b=>b.spread),backgroundColor:s2.map((_,i)=>CL[i%CL.length]),borderRadius:3}}]}},
    options:{{indexAxis:'y',responsive:true,maintainAspectRatio:false,
      plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:c=>c.parsed.x+' bps'}}}}}},
      scales:{{x:{{title:{{display:true,text:'G-spread (bps)'}},grid:{{color:'rgba(42,53,80,.3)'}}}},y:{{ticks:{{font:{{size:10}}}},grid:{{display:false}}}}}}
    }}
  }});
}}

const TC=[
  {{id:'label',n:'채권명',t:'t'}},{{id:'ytm',n:'YTM(%)',t:'n',f:v=>v?.toFixed(2)}},
  {{id:'spread',n:'G-spread',t:'n',f:v=>v?.toFixed(0),c:'sp'}},
  {{id:'chg1d',n:'1D',t:'c',f:v=>v!=null?(v>=0?'+':'')+v.toFixed(1):'-'}},
  {{id:'chg1w',n:'1W',t:'c',f:v=>v!=null?(v>=0?'+':'')+v.toFixed(1):'-'}},
  {{id:'chg1m',n:'1M',t:'c',f:v=>v!=null?(v>=0?'+':'')+v.toFixed(1):'-'}},
  {{id:'chg3m',n:'3M',t:'c',f:v=>v!=null?(v>=0?'+':'')+v.toFixed(1):'-'}},
  {{id:'duration',n:'듀레이션',t:'n',f:v=>v?.toFixed(1)}},
  {{id:'matBucket',n:'만기',t:'t'}},{{id:'sector',n:'섹터',t:'t'}},
  {{id:'issuer',n:'발행자',t:'t'}},{{id:'rating',n:'등급',t:'t'}},{{id:'region',n:'지역',t:'t'}},
  {{id:'_qty',n:'수량(M)',t:'q'}},
  {{id:'_pnl',n:'손익($)',t:'p'}},
];

// 수량 localStorage
function getPnLQty(){{try{{return JSON.parse(localStorage.getItem('pnl_qty')||'{{}}')}}catch{{return{{}}}}}}
function setPnLQty(d){{localStorage.setItem('pnl_qty',JSON.stringify(d))}}
function qtyKey(label){{return label.replace(/[^a-zA-Z0-9]/g,'_')}}

function getScenario(){{return document.getElementById('pnl-scenario').value}}
function getDelta(b){{
  const s=getScenario();
  if(s==='1d')return b.chg1d||0;
  if(s==='1w')return b.chg1w||0;
  if(s==='1m')return b.chg1m||0;
  if(s==='3m')return b.chg3m||0;
  return 0;
}}

function RH(){{
  var h='';
  for(var i=0;i<TC.length;i++){{
    var c=TC[i];
    var a=sc===c.id?(sd>0?'▲':'▼'):'';
    var cls=sc===c.id?'so':'';
    h+='<th class="'+cls+'" data-col="'+c.id+'">'+c.n+' <span style="font-size:10px">'+a+'</span></th>';
  }}
  document.getElementById('th').innerHTML=h;
  // Attach click handlers
  var ths=document.getElementById('th').querySelectorAll('th');
  for(var j=0;j<ths.length;j++){{
    ths[j].addEventListener('click',(function(col){{return function(){{SO(col)}}}})( TC[j].id ));
  }}
}}
function SO(c){{if(sc===c)sd*=-1;else{{sc=c;sd=-1}}RT()}}
function RT(){{
  RH();
  let d=GF2();
  const savedQty=getPnLQty();

  // _qty와 _pnl을 미리 계산해서 정렬 가능하게
  d.forEach(function(b){{
    const k=qtyKey(b.label);
    b._qty=savedQty[k]!==undefined?savedQty[k]:0;
    const delta=getDelta(b);
    b._pnl=(b.duration&&b._qty)?-b.duration*(delta/10000)*b._qty*1000000:0;
  }});

  d.sort(function(a,b){{const av=a[sc],bv=b[sc];if(av==null)return 1;if(bv==null)return -1;
    if(typeof av==='string')return av.localeCompare(bv)*sd;return(av-bv)*sd}});

  const rows=[];
  for(var j=0;j<d.length;j++){{
    var b=d[j];
    var isSel=sel.has(b.label);
    var rs=isSel?'background:rgba(59,130,246,.06)':'';
    var cells=[];
    for(var ci=0;ci<TC.length;ci++){{
      var c=TC[ci];
      if(c.t==='q'){{
        // 수량 입력 필드
        cells.push('<td class="nm"><input type="number" value="'+b._qty+'" step="0.5" min="0" style="width:60px;padding:3px 6px;font-size:11px;font-family:var(--mn);background:var(--sf2);color:var(--tx);border:1px solid var(--bd);border-radius:4px;text-align:right" data-key="'+qtyKey(b.label)+'" onchange="updateQty(this)" onclick="event.stopPropagation()"></td>');
      }} else if(c.t==='p'){{
        // P&L
        var pnl=b._pnl;
        if(!b._qty||b._qty===0){{
          cells.push('<td class="nm" style="color:var(--tx3)">-</td>');
        }} else {{
          var pnlFmt=pnl>=0?'+'+Math.round(pnl).toLocaleString():Math.round(pnl).toLocaleString();
          var pnlCls=pnl>0?'ng':pnl<0?'ps':'';
          cells.push('<td class="nm '+pnlCls+'" style="font-weight:600">'+pnlFmt+'</td>');
        }}
      }} else {{
        var v=b[c.id];
        var cls=c.t==='t'?'tc':'nm';
        if(c.c)cls+=' '+c.c;
        if(c.t==='c'&&v!=null)cls+=v>0?' ps':v<0?' ng':'';
        var display=c.f?c.f(v):(v!=null?v:'-');
        cells.push('<td class="'+cls+'">'+display+'</td>');
      }}
    }}
    rows.push('<tr style="'+rs+'" data-label="'+b.label.replace(/"/g,'&quot;')+'">' + cells.join('') + '</tr>');
  }}
  document.getElementById('tb').innerHTML=rows.join('');
  // Row click delegation
  var trs=document.getElementById('tb').querySelectorAll('tr');
  for(var ri=0;ri<trs.length;ri++){{
    trs[ri].addEventListener('click',function(){{
      var lbl=this.getAttribute('data-label');
      if(lbl)TB(lbl);
    }});
  }}
  UC();
  updatePnLSummary(d);
}}

function updatePnLSummary(data){{
  var totalPnL=0, totalNotional=0, cnt=0;
  for(var i=0;i<data.length;i++){{
    if(data[i]._qty>0){{
      totalPnL+=data[i]._pnl;
      totalNotional+=data[i]._qty;
      cnt++;
    }}
  }}
  if(cnt===0){{
    document.getElementById('pnl-summary').innerHTML='💰 수량을 입력하면 포트폴리오 손익이 자동 계산됩니다 (수량은 브라우저에 저장)';
    return;
  }}
  var totalFmt=totalPnL>=0?'+$'+Math.abs(Math.round(totalPnL)).toLocaleString():'-$'+Math.abs(Math.round(totalPnL)).toLocaleString();
  var totalCls=totalPnL>0?'color:var(--gn)':totalPnL<0?'color:var(--rd)':'';
  var scenarioLabel={{'1d':'1일','1w':'1주','1m':'1개월','3m':'3개월'}}[getScenario()]||getScenario();
  document.getElementById('pnl-summary').innerHTML=
    '<b style="'+totalCls+';font-size:16px">'+totalFmt+'</b>'
    +' <span style="color:var(--tx3);font-size:12px">| '+scenarioLabel+' 스프레드 변동 기준 | '
    +'보유 '+cnt+'개 | 총 액면 '+totalNotional.toFixed(1)+'M</span>';
}}

function updateQty(el){{
  var d=getPnLQty();
  d[el.getAttribute('data-key')]=parseFloat(el.value)||0;
  setPnLQty(d);
  RT();
}}

function TB(l){{if(sel.has(l))sel.delete(l);else sel.add(l);R1()}}

function PF(){{
  var ss=[...new Set(bonds.map(b=>b.sector).filter(Boolean))].sort();
  var ms=[...new Set(bonds.map(b=>b.matBucket).filter(Boolean))];
  var rs=[...new Set(bonds.map(b=>b.region).filter(Boolean))].sort();
  // Chart filters
  ss.forEach(function(s){{var o=document.createElement('option');o.value=s;o.textContent=s;document.getElementById('fs').appendChild(o)}});
  ms.forEach(function(m){{var o=document.createElement('option');o.value=m;o.textContent=m;document.getElementById('fm').appendChild(o)}});
  rs.forEach(function(r){{var o=document.createElement('option');o.value=r;o.textContent=r;document.getElementById('fr').appendChild(o)}});
  // Table filters
  ss.forEach(function(s){{var o=document.createElement('option');o.value=s;o.textContent=s;document.getElementById('fs2').appendChild(o)}});
  ms.forEach(function(m){{var o=document.createElement('option');o.value=m;o.textContent=m;document.getElementById('fm2').appendChild(o)}});
  rs.forEach(function(r){{var o=document.createElement('option');o.value=r;o.textContent=r;document.getElementById('fr2').appendChild(o)}});
}}

function R1(){{
  try{{RM()}}catch(e){{console.error('RM:',e)}}
  try{{RS()}}catch(e){{console.error('RS:',e)}}
  try{{RC1()}}catch(e){{console.error('RC1:',e)}}
  try{{RC2()}}catch(e){{console.error('RC2:',e)}}
  try{{RC3()}}catch(e){{console.error('RC3:',e)}}
  try{{RT()}}catch(e){{console.error('RT:',e)}}
}}

// ═══ TAB 2 ═══
let t2r=false;
function R2(){{
  if(t2r)return;t2r=true;
  const mc=D.marketCurrent,mg=D.marketChanges,mt=D.marketTimeseries;
  const names=Object.keys(mc);
  const fc=(v)=>v!=null?(+v>=0?'+':'')+Number(v).toFixed(1):'-';
  const cc=(v)=>v!=null?(+v>0?'up':+v<0?'dn':''):'';
  document.getElementById('m2').innerHTML=names.map(n=>{{
    const v=mc[n],g=mg[n]||{{}};const u=n.includes('VIX')||n.includes('MOVE')?'':' bps';
    return`<div class="mc"><div class="lb">${{n}}</div><div class="vl">${{v!=null?v.toFixed(1)+u:'-'}}</div>
      <div style="display:flex;gap:10px;font-size:11px;margin-top:6px;font-family:var(--mn)">
        <span class="dl ${{cc(g['1d'])}}">1D:${{fc(g['1d'])}}</span>
        <span class="dl ${{cc(g['1w'])}}">1W:${{fc(g['1w'])}}</span>
        <span class="dl ${{cc(g['1m'])}}">1M:${{fc(g['1m'])}}</span></div></div>`;
  }}).join('');

  const cnames=['US IG OAS','US HY OAS','CDX IG 5Y','CDX HY 5Y'];
  const mkds=(n)=>{{const ts=mt[n];if(!ts)return{{dates:[],vals:[]}};const k=Object.keys(ts).sort();return{{dates:k,vals:k.map(d=>ts[d])}}}};
  const cd0=mkds(cnames[0]);
  new Chart(document.getElementById('c4').getContext('2d'),{{type:'line',
    data:{{labels:cd0.dates,datasets:cnames.map((n,i)=>{{const d=mkds(n);return{{label:n,data:d.vals,borderColor:CL[i],borderWidth:1.8,pointRadius:0,tension:.3,yAxisID:n.includes('HY')?'y1':'y'}}}})
    }},options:{{responsive:true,maintainAspectRatio:false,interaction:{{mode:'index',intersect:false}},
      scales:{{x:{{ticks:{{maxTicksLimit:8}},grid:{{display:false}}}},y:{{title:{{display:true,text:'IG (bps)'}},position:'left',grid:{{color:'rgba(42,53,80,.3)'}}}},y1:{{title:{{display:true,text:'HY (bps)'}},position:'right',grid:{{display:false}}}}}}
    }}
  }});

  const vd=mkds('VIX');
  new Chart(document.getElementById('c5').getContext('2d'),{{type:'line',
    data:{{labels:vd.dates,datasets:[{{label:'VIX',data:vd.vals,borderColor:'#ef4444',borderWidth:1.5,pointRadius:0,tension:.3,fill:{{target:'origin',above:'rgba(239,68,68,.08)'}}}}]}},
    options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},scales:{{x:{{ticks:{{maxTicksLimit:6}},grid:{{display:false}}}},y:{{grid:{{color:'rgba(42,53,80,.3)'}}}}}}}}
  }});

  const md=mkds('MOVE');
  new Chart(document.getElementById('c6').getContext('2d'),{{type:'line',
    data:{{labels:md.dates,datasets:[{{label:'MOVE',data:md.vals,borderColor:'#f59e0b',borderWidth:1.5,pointRadius:0,tension:.3,fill:{{target:'origin',above:'rgba(245,158,11,.08)'}}}}]}},
    options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{display:false}}}},scales:{{x:{{ticks:{{maxTicksLimit:6}},grid:{{display:false}}}},y:{{grid:{{color:'rgba(42,53,80,.3)'}}}}}}}}
  }});

  // ETF 스프레드 테이블
  const et=D.etfTimeseries||{{}};
  const etNames=Object.keys(D.etfCurrent||{{}});
  const emkds=(n)=>{{const ts=et[n];if(!ts)return{{dates:[],vals:[]}};const k=Object.keys(ts).sort();return{{dates:k,vals:k.map(d=>ts[d])}}}};

  if(etNames.length){{
    const ec=D.etfCurrent||{{}};
    const eg=D.etfChanges||{{}};
    const fc2=(v)=>v!=null?(+v>=0?'+':'')+Number(v).toFixed(1):'-';
    const cc2=(v)=>v!=null?(+v>0?'ps':+v<0?'ng':''):'';
    document.getElementById('etfTbl').innerHTML=etNames.map(n=>{{
      const g=eg[n]||{{}};
      return`<tr>
        <td class="tc" style="font-weight:600">${{n}}</td>
        <td class="nm sp">${{ec[n]!=null?ec[n].toFixed(0):'-'}}</td>
        <td class="nm ${{cc2(g['1d'])}}">${{fc2(g['1d'])}}</td>
        <td class="nm ${{cc2(g['1w'])}}">${{fc2(g['1w'])}}</td>
        <td class="nm ${{cc2(g['1m'])}}">${{fc2(g['1m'])}}</td>
        <td class="nm ${{cc2(g['3m'])}}">${{fc2(g['3m'])}}</td>
        <td class="nm ${{cc2(g['6m'])}}">${{fc2(g['6m'])}}</td>
      </tr>`;
    }}).join('');

    const eCL=['#3b82f6','#10b981','#8b5cf6','#06b6d4','#6366f1','#ef4444','#f59e0b','#ec4899'];
    const igETFs=['VCSH','VCIT','LQD','VCLT','IEAC'];
    const hyETFs=['HYG','EMB','EMLC'];

    // IG ETFs — dual axis: VCSH left (tight spread), VCLT right (wide spread)
    const igData=igETFs.filter(n=>et[n]);
    if(igData.length){{
      const igd0=emkds(igData[0]);
      const leftETFs=['VCSH','VCIT'];
      const rightETFs=['LQD','VCLT','IEAC'];
      new Chart(document.getElementById('c9').getContext('2d'),{{type:'line',
        data:{{labels:igd0.dates,datasets:igData.map((n,i)=>{{
          const d=emkds(n);
          const isRight=rightETFs.includes(n);
          return{{label:n,data:d.vals,borderColor:eCL[i%eCL.length],borderWidth:1.8,pointRadius:0,tension:.3,yAxisID:isRight?'y1':'y'}}
        }})
        }},options:{{responsive:true,maintainAspectRatio:false,interaction:{{mode:'index',intersect:false}},
          plugins:{{legend:{{position:'top'}},tooltip:{{callbacks:{{label:c=>c.dataset.label+': '+c.parsed.y?.toFixed(0)+' bps'}}}}}},
          scales:{{
            x:{{ticks:{{maxTicksLimit:6}},grid:{{display:false}}}},
            y:{{title:{{display:true,text:'VCSH/VCIT (bps)'}},position:'left',grid:{{color:'rgba(42,53,80,.3)'}}}},
            y1:{{title:{{display:true,text:'LQD/VCLT/IEAC (bps)'}},position:'right',grid:{{display:false}}}}
          }}
        }}
      }});
    }}

    // HY/EM ETFs
    const hyData=hyETFs.filter(n=>et[n]);
    if(hyData.length){{
      const hyd0=emkds(hyData[0]);
      new Chart(document.getElementById('c10').getContext('2d'),{{type:'line',
        data:{{labels:hyd0.dates,datasets:hyData.map((n,i)=>{{const d=emkds(n);return{{label:n,data:d.vals,borderColor:eCL[(i+5)%eCL.length],borderWidth:1.8,pointRadius:0,tension:.3}}}})
        }},options:{{responsive:true,maintainAspectRatio:false,interaction:{{mode:'index',intersect:false}},
          plugins:{{legend:{{position:'top'}},tooltip:{{callbacks:{{label:c=>c.dataset.label+': '+c.parsed.y?.toFixed(0)+' bps'}}}}}},
          scales:{{x:{{ticks:{{maxTicksLimit:6}},grid:{{display:false}}}},y:{{title:{{display:true,text:'OAS (bps)'}},grid:{{color:'rgba(42,53,80,.3)'}}}}}}
        }}
      }});
    }}

  // ─── BDC 카드 + 차트 ───────────────────────────────────────────
  const bdcC = D.bdcCurrent   || {{}};
  const bdcG = D.bdcChanges   || {{}};
  const bdcT = D.bdcTimeseries|| {{}};
  const bdcM = D.bdcMeta      || {{}};
  const bdcNames = Object.keys(bdcM);

  if(bdcNames.length > 0) {{
    const fc3  = (v) => v!=null ? (Number(v)>=0?'+':'')+Number(v).toFixed(1)+'%' : '—';
    const cc3  = (v) => v!=null ? (Number(v)>0?'up':Number(v)<0?'dn':'') : '';
    const fmtPB= (pb) => {{
      if(pb==null) return '—';
      const disc=(pb-1)*100;
      const col = disc<-20?'#dc2626': disc<-10?'#D35400': disc<0?'#f59e0b':'#16a34a';
      const sign= disc>=0?'+':'';
      return '<b style="color:'+col+'">'+pb.toFixed(2)+'x</b>'
            +'<span style="font-size:10px;color:'+col+'"> ('+sign+disc.toFixed(1)+'%)</span>';
    }};

    // 카드
    document.getElementById('bdcCards').innerHTML = bdcNames.map(nm => {{
      const cur=bdcC[nm]||{{}}, chg=bdcG[nm]||{{}}, meta=bdcM[nm]||{{}};
      const px=cur.price, pb=cur.pb;
      return '<div style="flex:1;min-width:130px;background:var(--bg1);border:1px solid var(--bd);'
            +'border-top:3px solid '+meta.color+';border-radius:6px;padding:9px 11px;">'
            +'<div style="font-size:10px;color:var(--tx3);margin-bottom:1px">'+meta.name+'</div>'
            +'<div style="font-size:13px;font-weight:700;color:var(--tx1)">'+nm
            +(px?' <span style="font-weight:400;font-size:11px">$'+px.toFixed(2)+'</span>':'')+'</div>'
            +'<div style="font-size:11px;margin:3px 0">P/BV: '+fmtPB(pb)+'</div>'
            +'<div style="font-size:10px;display:flex;gap:5px;">'
            +'<span class="dl '+cc3(chg['1d'])+'">1D:'+fc3(chg['1d'])+'</span>'
            +'<span class="dl '+cc3(chg['1w'])+'">1W:'+fc3(chg['1w'])+'</span>'
            +'<span class="dl '+cc3(chg['1m'])+'">1M:'+fc3(chg['1m'])+'</span>'
            +'</div></div>';
    }}).join('');

    // NAV 할인율 바 차트
    const bL=[], bV=[], bColors=[];
    bdcNames.forEach(nm => {{
      const pb=(bdcC[nm]||{{}}).pb;
      if(pb!=null) {{
        const disc=parseFloat(((pb-1)*100).toFixed(1));
        bL.push(nm); bV.push(disc);
        bColors.push(disc<-20?'#dc2626': disc<-10?'#D35400': disc<0?'#f59e0b':'#16a34a');
      }}
    }});
    if(bL.length > 0) {{
      new Chart(document.getElementById('bdcBarChart').getContext('2d'), {{
        type:'bar',
        data:{{labels:bL, datasets:[{{label:'NAV 할인율(%)',data:bV,
          backgroundColor:bColors.map(c=>c+'BB'), borderColor:bColors,
          borderWidth:1.5, borderRadius:4}}]}},
        options:{{responsive:true,maintainAspectRatio:false,
          plugins:{{legend:{{display:false}},
            tooltip:{{callbacks:{{label:c=>{{const v=c.parsed.y;return(v>=0?'+':'')+v.toFixed(1)+'% ('+(v>=0?'프리미엄':'할인')+')';}}}},
          scales:{{
            x:{{grid:{{display:false}},ticks:{{font:{{size:11}}}},
            y:{{grid:{{color:'rgba(42,53,80,.3)'}},ticks:{{font:{{size:10}},callback:v=>v+'%'}}
          }}
      }});
    }}

    // 주가 시계열 (Indexed=100)
    const palette=['#1B3A6B','#2E5FA3','#D35400','#1A6B3C','#C0392B','#8E44AD'];
    const tDatasets=[];
    let tLabels=[];
    bdcNames.forEach((nm,i) => {{
      const ts=bdcT[nm];
      if(!ts) return;
      const dates=Object.keys(ts).sort();
      const vals=dates.map(d=>ts[d]);
      if(vals.length<5 || vals[0]===0) return;
      const base=vals[0];
      const indexed=vals.map(v=>parseFloat((v/base*100).toFixed(2)));
      if(dates.length > tLabels.length) tLabels=dates;
      tDatasets.push({{label:nm,data:indexed,
        borderColor:palette[i%palette.length],
        backgroundColor:'transparent',
        borderWidth:1.8,pointRadius:0,tension:.2}});
    }});
    if(tDatasets.length > 0) {{
      new Chart(document.getElementById('bdcLineChart').getContext('2d'), {{
        type:'line',
        data:{{labels:tLabels, datasets:tDatasets}},
        options:{{responsive:true,maintainAspectRatio:false,
          interaction:{{mode:'index',intersect:false}},
          plugins:{{legend:{{position:'top',labels:{{font:{{size:10}},boxWidth:12,padding:6}}}},
          scales:{{
            x:{{grid:{{display:false}},ticks:{{maxTicksLimit:6,font:{{size:10}},maxRotation:0}},
            y:{{grid:{{color:'rgba(42,53,80,.3)'}},ticks:{{font:{{size:10}}}}
          }}
      }});
    }}
  }}

  }}
}}

// ═══ TAB 3: Issuer Curves (fixed) ═══
let t3r=false;
function R3(){{
  if(t3r)return;t3r=true;
  const all=D.allBonds;

  // Group by issuer
  const byIssuer={{}};
  all.forEach(b=>{{if(b.issuer&&b.duration&&b.ytm){{(byIssuer[b.issuer]=byIssuer[b.issuer]||[]).push(b)}} }});

  // Top issuers by bond count (min 3 bonds)
  const top=Object.entries(byIssuer).filter(([,bs])=>bs.length>=3).sort((a,b)=>b[1].length-a[1].length).slice(0,10);

  const container=document.getElementById('ic');
  container.innerHTML=top.map(([issuer,bs],idx)=>{{
    const owned=bs.filter(b=>b.is_owned).length;
    const extra=bs.filter(b=>!b.is_owned).length;
    const sub=`보유 ${{owned}}개`+(extra>0?` + 비보유 ${{extra}}개`:'');
    return`<div class="ic">
      <div class="it">${{issuer}}</div>
      <div class="is">${{sub}} — <span style="color:var(--ac)">● 현재</span> <span style="color:var(--am)">◆ 1M전</span> <span style="color:var(--rd)">▲ 3M전</span></div>
      <div class="ch" style="height:260px"><canvas id="ic${{idx}}"></canvas></div></div>`;
  }}).join('');

  top.forEach(([issuer,bs],idx)=>{{
    const color=CL[idx%CL.length];
    const allSorted=[...bs].sort((a,b)=>a.duration-b.duration);

    const datasets=[];

    // Current YTM — all bonds as line
    datasets.push({{
      label:'현재',
      data:allSorted.map(b=>({{x:b.duration,y:b.ytm}})),
      borderColor:color,backgroundColor:color,
      pointRadius:allSorted.map(b=>b.is_owned?7:5),
      pointStyle:allSorted.map(b=>b.is_owned?'circle':'circle'),
      pointBackgroundColor:allSorted.map(b=>b.is_owned?color:'transparent'),
      pointBorderColor:color,
      pointBorderWidth:allSorted.map(b=>b.is_owned?0:2),
      pointHoverRadius:9,
      showLine:true,tension:.3,borderWidth:2,
    }});

    // 1M ago YTM
    const has1m=allSorted.filter(b=>b.ytm_1m_ago!=null);
    if(has1m.length>=2){{
      datasets.push({{
        label:'1M전',
        data:has1m.map(b=>({{x:b.duration,y:b.ytm_1m_ago}})),
        borderColor:'#f59e0b',backgroundColor:'transparent',
        pointRadius:4,pointStyle:'rect',pointBorderColor:'#f59e0b',pointBorderWidth:1.5,
        pointHoverRadius:6,showLine:true,tension:.3,borderWidth:1.5,borderDash:[4,3],
      }});
    }}

    // 3M ago YTM
    const has3m=allSorted.filter(b=>b.ytm_3m_ago!=null);
    if(has3m.length>=2){{
      datasets.push({{
        label:'3M전',
        data:has3m.map(b=>({{x:b.duration,y:b.ytm_3m_ago}})),
        borderColor:'#ef4444',backgroundColor:'transparent',
        pointRadius:4,pointStyle:'triangle',pointBorderColor:'#ef4444',pointBorderWidth:1.5,
        pointHoverRadius:6,showLine:true,tension:.3,borderWidth:1.5,borderDash:[2,2],
      }});
    }}

    const ctx=document.getElementById('ic'+idx).getContext('2d');
    new Chart(ctx,{{type:'scatter',data:{{datasets}},
      options:{{responsive:true,maintainAspectRatio:false,
        plugins:{{
          legend:{{display:true,position:'top',labels:{{font:{{size:10}},padding:8}}}},
          tooltip:{{callbacks:{{
            label:function(c){{
              const b=c.datasetIndex===0?allSorted[c.dataIndex]:(c.datasetIndex===1?has1m[c.dataIndex]:has3m[c.dataIndex]);
              if(!b)return c.dataset.label+': '+c.parsed.y?.toFixed(2)+'%';
              return b.label+' | '+c.parsed.x.toFixed(1)+'Y | '+c.parsed.y.toFixed(2)+'%';
            }}
          }}}}
        }},
        scales:{{
          x:{{title:{{display:true,text:'듀레이션(년)',font:{{size:10}}}},
              grid:{{color:'rgba(42,53,80,.2)'}},
              ticks:{{callback:v=>[1,2,3,5,7,10,15,20,30].includes(Math.round(v))?Math.round(v):'',font:{{size:10}}}}}},
          y:{{title:{{display:true,text:'YTM(%)',font:{{size:10}}}},
              ticks:{{callback:v=>v.toFixed(1)+'%',font:{{size:10}}}},
              grid:{{color:'rgba(42,53,80,.2)'}}}}
        }}
      }}
    }});
  }});

  // Sector average curve — current only
  const bySector={{}};
  all.forEach(b=>{{if(b.sector&&b.duration&&b.ytm)(bySector[b.sector]=bySector[b.sector]||[]).push(b)}});
  const sds=Object.entries(bySector).filter(([,bs])=>bs.length>=3).map(([sec,bs],i)=>{{
    const bk={{}};bs.forEach(b=>{{const k=Math.round(b.duration);(bk[k]=bk[k]||[]).push(b.ytm)}});
    const pts=Object.entries(bk).map(([d,ys])=>({{x:+d,y:ys.reduce((a,v)=>a+v,0)/ys.length}})).sort((a,b)=>a.x-b.x);
    return{{label:sec,data:pts,borderColor:CL[i%CL.length],borderWidth:2,pointRadius:4,showLine:true,tension:.3}};
  }});
  new Chart(document.getElementById('c7').getContext('2d'),{{type:'scatter',data:{{datasets:sds}},
    options:{{responsive:true,maintainAspectRatio:false,plugins:{{legend:{{position:'top'}}}},
      scales:{{x:{{title:{{display:true,text:'듀레이션(년)'}},grid:{{color:'rgba(42,53,80,.3)'}}}},y:{{title:{{display:true,text:'YTM(%)'}},grid:{{color:'rgba(42,53,80,.3)'}}}}}}
    }}
  }});
}}

// ═══ TAB 4 ═══
let t4r=false;
function R4(){{
  if(t4r)return;t4r=true;
  const tr=D.trades||[];
  var rows=[];
  for(var i=0;i<tr.length;i++){{
    var t=tr[i];
    var spChgCls=t.spread_chg!=null?(t.spread_chg>0?'ps':t.spread_chg<0?'ng':''):'';
    var spChgFmt=t.spread_chg!=null?((t.spread_chg>=0?'+':'')+t.spread_chg.toFixed(1)):'-';
    rows.push('<tr>'
      +'<td class="tc">'+( t.date||'-')+'</td>'
      +'<td class="tc">'+(t.bond||'-')+'</td>'
      +'<td class="tc" style="font-size:10px">'+(t.source||'-')+'</td>'
      +'<td class="nm">'+(t.price!=null?t.price:'-')+'</td>'
      +'<td class="nm">'+(t.ytm!=null?t.ytm:'-')+'</td>'
      +'<td class="nm sp">'+(t.trd_spread!=null?t.trd_spread.toFixed(1):'-')+'</td>'
      +'<td class="nm">'+(t.yas_spread!=null?t.yas_spread.toFixed(1):'-')+'</td>'
      +'<td class="nm '+spChgCls+'" style="font-weight:600">'+spChgFmt+'</td>'
      +'<td class="nm">'+(t.volume?t.volume.toLocaleString():'-')+'</td>'
      +'<td class="nm">'+(t.volume?(t.volume/1e6).toFixed(1)+'M':'-')+'</td>'
      +'</tr>');
  }}
  document.getElementById('t4b').innerHTML=rows.join('');
}}

// ═══ TAB 6: New Issues ═══
let t6r=false;
function R6(){{
  try{{
    if(!t6r){{
      t6r=true;
      const cors=[...new Set((D.newIssues||[]).map(i=>i.cor).filter(Boolean))].sort();
      cors.forEach(c=>{{const o=document.createElement('option');o.value=c;o.textContent=c;document.getElementById('nf3').appendChild(o)}});
    }}
    RNI();
  }}catch(e){{document.getElementById('niCount').textContent='Error: '+e.message;console.error(e)}}
}}
function RNI(){{
  const ni=D.newIssues||[];
  if(!ni.length){{document.getElementById('niCount').textContent='데이터 없음';document.getElementById('nib').innerHTML='<tr><td colspan="15" style="text-align:center;padding:20px;color:var(--tx3)">발행 데이터가 없습니다</td></tr>';return}}
  const f1=document.getElementById('nf1').value;
  const f2=document.getElementById('nf2').value;
  const f3=document.getElementById('nf3').value;
  const f4=document.getElementById('nf4').value.toLowerCase();
  const filtered=ni.filter(i=>{{
    if(f1!=='ALL'&&i.ighy!==f1)return false;
    if(f2!=='ALL'&&i.status!==f2)return false;
    if(f3!=='ALL'&&i.cor!==f3)return false;
    if(f4&&!(i.issuer||'').toLowerCase().includes(f4)&&!(i.ticker||'').toLowerCase().includes(f4))return false;
    return true;
  }});
  document.getElementById('niCount').textContent=filtered.length+'건 표시 / 전체 '+ni.length+'건';
  const sc2=function(s){{if(!s)return'';if(s==='Priced')return'color:var(--gn)';if(s==='Talk'||s==='Launch')return'color:var(--am)';return''}};
  const igc=function(v){{if(!v)return'';if(v==='IG')return'color:var(--ac)';if(v==='HY')return'color:var(--rd)';return''}};
  const fmtNic=function(n){{if(n===null||n===undefined)return'-';try{{return Number(n).toFixed(1)}}catch(e){{return'-'}}}};
  const fmtCpn=function(c){{if(c===null||c===undefined)return'-';try{{return Number(c).toFixed(3)}}catch(e){{return'-'}}}};
  const nicCls=function(n){{if(n===null||n===undefined)return'';return n>0?'ng':'ps'}};
  const rows=[];
  for(let j=0;j<filtered.length;j++){{
    const i=filtered[j];
    rows.push('<tr>'
      +'<td class="tc">'+(i.date||'-')+'</td>'
      +'<td class="tc">'+(i.cor||'-')+'</td>'
      +'<td class="tc" style="font-weight:600;'+igc(i.ighy)+'">'+(i.ighy||'-')+'</td>'
      +'<td class="tc" style="font-size:11px">'+(i.ticker||'-')+'</td>'
      +'<td class="tc">'+(i.issuer||'-')+'</td>'
      +'<td class="nm" style="font-size:11px">'+(i.ipt||'-')+'</td>'
      +'<td class="nm" style="font-size:11px;font-weight:600">'+(i.fg||'-')+'</td>'
      +'<td class="nm '+nicCls(i.nic)+'">'+fmtNic(i.nic)+'</td>'
      +'<td class="nm">'+(i.amt||'-')+'</td>'
      +'<td class="tc">'+(i.ccy||'-')+'</td>'
      +'<td class="nm">'+fmtCpn(i.coupon)+'</td>'
      +'<td class="tc">'+(i.tenor||'-')+'</td>'
      +'<td class="tc" style="font-weight:600;'+sc2(i.status)+'">'+(i.status||'-')+'</td>'
      +'<td class="tc" style="font-size:10px">'+(i.rank||'-')+'</td>'
      +'<td class="tc" style="font-size:10px">'+(i.reg||'-')+'</td>'
      +'</tr>');
  }}
  document.getElementById('nib').innerHTML=rows.join('');
}}

// ═══ TAB 5 ═══
function GM(){{try{{return JSON.parse(localStorage.getItem('credit_memos')||'[]')}}catch{{return[]}}}}
function WM(a){{localStorage.setItem('credit_memos',JSON.stringify(a))}}
function SM(){{
  const k=document.getElementById('rk').value.trim(),t=document.getElementById('rt').value.trim(),c=document.getElementById('rc').value.trim();
  if(!k||!t||!c){{document.getElementById('rs').textContent='⚠️ 모든 필드를 입력해주세요';return}}
  const ms=GM();ms.push({{key:k,title:t,content:c,date:new Date().toISOString().split('T')[0],saved:new Date().toLocaleString('ko-KR')}});
  WM(ms);document.getElementById('rk').value='';document.getElementById('rt').value='';document.getElementById('rc').value='';
  document.getElementById('rs').textContent='✅ 저장됨';RL();
}}
function DM(i){{const ms=GM();ms.splice(i,1);WM(ms);RL()}}
function RL(){{
  const ms=GM();
  if(!ms.length){{document.getElementById('rl').innerHTML='<div style="text-align:center;padding:20px;color:var(--tx3);font-size:13px">저장된 메모가 없어요</div>';return}}
  document.getElementById('rl').innerHTML=[...ms].reverse().map((m,ri)=>{{
    const i=ms.length-1-ri;
    return`<div class="mi"><div class="mh"><div><span class="mt">${{m.key}}</span><span class="mtl">${{m.title}}</span></div>
      <div style="display:flex;align-items:center;gap:10px"><span class="md">${{m.date}}</span>
        <button class="bt" style="padding:2px 8px;font-size:11px" onclick="DM(${{i}})">삭제</button></div></div>
      <div class="mb">${{m.content}}</div></div>`;
  }}).join('');
}}

// ═══ CSV Export ═══
function EX(){{
  const hdr=TC.map(c=>c.n).join(',');
  const rows=bonds.map(b=>TC.map(c=>{{let v=b[c.id];if(typeof v==='string'&&v.includes(','))return`"${{v}}"`;return v??''}}).join(','));
  const csv='\\uFEFF'+hdr+'\\n'+rows.join('\\n');
  const blob=new Blob([csv],{{type:'text/csv;charset=utf-8'}});
  const u=URL.createObjectURL(blob);const a=document.createElement('a');
  a.href=u;a.download=`ig_credit_${{D.generated.split(' ')[0]}}.csv`;a.click();URL.revokeObjectURL(u);
  const e=document.getElementById('en');e.classList.add('show');setTimeout(()=>e.classList.remove('show'),2000);
}}

// ═══ TAB 7: Bond Management ═══
var bondList=(D.bondIsins||[]).slice();

function R7(){{
  renderBondList();
}}

function renderBondList(){{
  document.getElementById('bondListCount').textContent=bondList.length;
  // Build isin lookup from all bonds data
  var lookup={{}};
  var allB=D.bonds.concat(D.allBonds||[]);
  for(var j=0;j<allB.length;j++){{
    if(allB[j].isin)lookup[allB[j].isin]=allB[j];
  }}
  var rows=[];
  for(var i=0;i<bondList.length;i++){{
    var isin=bondList[i];
    var b=lookup[isin]||null;
    rows.push('<tr>'
      +'<td style="text-align:center"><input type="checkbox" class="bl-chk" data-idx="'+i+'"></td>'
      +'<td class="tc" style="font-family:var(--mn);font-size:11px">'+isin+'</td>'
      +'<td class="tc">'+(b&&b.label?b.label:'-')+'</td>'
      +'<td class="tc">'+(b&&b.issuer?b.issuer:'-')+'</td>'
      +'<td class="tc">'+(b&&b.sector?b.sector:'-')+'</td>'
      +'</tr>');
  }}
  document.getElementById('blBody').innerHTML=rows.join('');
}}

function addBond(){{
  var input=document.getElementById('addIsin').value.trim();
  if(!input){{document.getElementById('addMsg').textContent='⚠️ ISIN을 입력해주세요';return}}
  // 여러개 입력 지원 (쉼표/공백/줄바꿈 구분)
  var items=input.split(/[,\s]+/).filter(function(s){{return s.length>0}});
  var added=0;
  for(var i=0;i<items.length;i++){{
    var isin=items[i].replace(/\s+Corp$/i,'').trim();
    if(isin&&bondList.indexOf(isin)<0){{
      bondList.push(isin);
      added++;
    }}
  }}
  document.getElementById('addIsin').value='';
  document.getElementById('addMsg').textContent='✅ '+added+'개 추가됨 (총 '+bondList.length+'개)';
  renderBondList();
}}

function selectAllBonds(){{
  var chks=document.querySelectorAll('.bl-chk');
  for(var i=0;i<chks.length;i++)chks[i].checked=true;
}}
function deselectAllBonds(){{
  var chks=document.querySelectorAll('.bl-chk');
  for(var i=0;i<chks.length;i++)chks[i].checked=false;
}}

function deleteSelected(){{
  var chks=document.querySelectorAll('.bl-chk:checked');
  if(!chks.length)return;
  var idxs=[];
  for(var i=0;i<chks.length;i++)idxs.push(parseInt(chks[i].getAttribute('data-idx')));
  idxs.sort(function(a,b){{return b-a}});
  for(var i=0;i<idxs.length;i++)bondList.splice(idxs[i],1);
  document.getElementById('addMsg').textContent='🗑 '+idxs.length+'개 삭제됨 (총 '+bondList.length+'개)';
  renderBondList();
}}

function downloadBondList(){{
  var json=JSON.stringify(bondList,null,2);
  var blob=new Blob([json],{{type:'application/json'}});
  var url=URL.createObjectURL(blob);
  var a=document.createElement('a');
  a.href=url;a.download='bond_list.json';a.click();
  URL.revokeObjectURL(url);
  document.getElementById('dlMsg').textContent='✅ 다운로드 완료! credit-dashboard 폴더에 저장하세요';
  setTimeout(function(){{document.getElementById('dlMsg').textContent=''}},5000);
}}

// ═══ INIT ═══
PF();R1();
</script>
</body>
</html>'''
    return html


# ══════════════════════════════════════════════
# 6. GitHub Pages 자동 배포
# ══════════════════════════════════════════════
def deploy_to_github(html_content, repo_dir=None):
    """
    GitHub Pages 저장소에 index.html을 업데이트하고 push.
    repo_dir: git 저장소 경로. None이면 스크립트와 같은 폴더 사용.
    """
    import subprocess

    if repo_dir is None:
        repo_dir = os.path.dirname(os.path.abspath(__file__))

    index_path = os.path.join(repo_dir, "index.html")

    # index.html 저장 (항상 같은 파일 → 같은 URL)
    with open(index_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    today_str = datetime.today().strftime("%Y-%m-%d %H:%M")

    try:
        def run(cmd):
            r = subprocess.run(cmd, cwd=repo_dir, capture_output=True, text=True, timeout=30)
            if r.returncode != 0:
                print(f"  ⚠️  {' '.join(cmd)}: {r.stderr.strip()}")
            return r.returncode == 0

        # git 상태 확인
        result = subprocess.run(["git", "status"], cwd=repo_dir, capture_output=True, text=True)
        if result.returncode != 0:
            print("\n  ⚠️  이 폴더가 git 저장소가 아닙니다.")
            print("  아래 초기 설정을 먼저 진행해주세요:\n")
            print_setup_guide()
            return False

        run(["git", "add", "index.html"])
        # 발행 누적 데이터도 push
        history_file = os.path.join(repo_dir, "new_issues_history.json")
        if os.path.exists(history_file):
            run(["git", "add", "new_issues_history.json"])
        # 종목 관리 파일도 push
        bond_list_file = os.path.join(repo_dir, "bond_list.json")
        if os.path.exists(bond_list_file):
            run(["git", "add", "bond_list.json"])
        run(["git", "commit", "-m", f"Update dashboard {today_str}"])

        print("\n  📤 GitHub에 push 중...")
        if run(["git", "push"]):
            print("  ✅ GitHub Pages 배포 완료!")
            # remote URL에서 페이지 주소 추출
            r = subprocess.run(["git", "remote", "get-url", "origin"],
                               cwd=repo_dir, capture_output=True, text=True)
            if r.returncode == 0:
                url = r.stdout.strip()
                # git@github.com:user/repo.git → user/repo
                # https://github.com/user/repo.git → user/repo
                if "github.com" in url:
                    parts = url.replace(".git","").split("github.com")[-1]
                    parts = parts.lstrip(":/")
                    user, repo = parts.split("/")[:2]
                    print(f"\n  🌐 https://{user}.github.io/{repo}/")
            return True
        else:
            print("  ⚠️  push 실패. 네트워크 또는 인증을 확인해주세요.")
            return False

    except FileNotFoundError:
        print("\n  ⚠️  git이 설치되어 있지 않습니다. git을 먼저 설치해주세요.")
        return False
    except Exception as e:
        print(f"\n  ⚠️  배포 오류: {e}")
        return False


def print_setup_guide():
    """초기 GitHub Pages 설정 가이드 출력"""
    print("""
  ┌─────────────────────────────────────────────────┐
  │          GitHub Pages 초기 설정 가이드           │
  ├─────────────────────────────────────────────────┤
  │                                                 │
  │  1. GitHub에서 새 저장소 만들기                  │
  │     → https://github.com/new                    │
  │     → 이름: credit-dashboard                    │
  │     → Public 선택                               │
  │                                                 │
  │  2. 이 폴더에서 git 초기화                      │
  │     cd [이 스크립트가 있는 폴더]                 │
  │     git init                                    │
  │     git remote add origin                       │
  │       https://github.com/[유저명]/credit-dashboard.git │
  │     git branch -M main                          │
  │     git add .                                   │
  │     git commit -m "Initial commit"              │
  │     git push -u origin main                     │
  │                                                 │
  │  3. GitHub Pages 활성화                         │
  │     → 저장소 Settings → Pages                   │
  │     → Source: Deploy from a branch              │
  │     → Branch: main / (root)                     │
  │     → Save                                      │
  │                                                 │
  │  4. 1-2분 후 접속 가능:                         │
  │     https://[유저명].github.io/credit-dashboard/ │
  │                                                 │
  │  5. 이후 이 스크립트만 실행하면 자동 업데이트   │
  │     python export_credit_dashboard.py           │
  │                                                 │
  └─────────────────────────────────────────────────┘
""")


# ══════════════════════════════════════════════
# 7. 메인 실행
# ══════════════════════════════════════════════
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Credit Dashboard → HTML + GitHub Pages 배포")
    parser.add_argument("--no-deploy", action="store_true", help="HTML만 생성, GitHub push 안 함")
    parser.add_argument("--setup", action="store_true", help="초기 설정 가이드 출력")
    parser.add_argument("--sample", action="store_true", help="샘플 데이터로 생성 (Bloomberg 무시)")
    args = parser.parse_args()

    if args.setup:
        print_setup_guide()
        exit(0)

    print("\n" + "=" * 60)
    print("  Credit Dashboard → HTML 생성기 + GitHub Pages 배포")
    print("=" * 60)

    # Bloomberg 연결 시도
    if args.sample:
        print("\n📊 샘플 데이터 모드")
        data = load_sample_data()
    else:
        try:
            from xbbg import blp
            print("\n✅ Bloomberg 연결 확인됨 → 실제 데이터 로드")
            data = load_bloomberg_data()
        except ImportError:
            print("\n⚠️  xbbg 미설치 → 샘플 데이터로 생성")
            data = load_sample_data()
        except Exception as e:
            print(f"\n⚠️  Bloomberg 연결 실패 ({e}) → 샘플 데이터로 생성")
            data = load_sample_data()

    # HTML 생성
    html_content = generate_html(data)

    # index.html로 저장 (GitHub Pages용 고정 파일명)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    index_path = os.path.join(script_dir, "index.html")
    with open(index_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    # 날짜별 백업도 저장
    backup_name = f"credit_dashboard_{datetime.today().strftime('%Y%m%d')}.html"
    backup_path = os.path.join(script_dir, backup_name)
    with open(backup_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    # JSON 데이터 백업
    json_name = f"credit_data_{datetime.today().strftime('%Y%m%d')}.json"
    json_path = os.path.join(script_dir, json_name)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    size_kb = os.path.getsize(index_path) / 1024
    print(f"\n  📄 index.html 생성 완료 ({size_kb:.0f} KB)")
    print(f"  📄 {backup_name} (날짜별 백업)")
    print(f"  📊 {json_name} (데이터 백업)")

    # GitHub Pages 배포
    if args.no_deploy:
        print(f"\n  --no-deploy 모드: GitHub push를 건너뜁니다.")
    else:
        deploy_to_github(html_content, script_dir)

    print(f"\n{'=' * 60}\n")
